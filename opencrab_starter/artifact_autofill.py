from __future__ import annotations

import json
import re
from copy import copy
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as WorksheetImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


STYLE_PATTERN = re.compile(r"(?<!\d)(\d{9})(?!\d)")
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".bmp"}

FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "style": ("style", "style no", "style number", "style #", "outlet style", "outlet style #"),
    "core_style": ("core style", "ref style", "reference style"),
    "description": ("description", "style description", "desc"),
    "sketch": ("sketch", "tp photo", "tp photos", "photo", "image", "image path", "preview path"),
    "fabric": ("fabric", "fabric information", "fabric info", "quality", "yarn"),
    "fabric_cost": ("fabric cost", "fabric price", "cif yd", "cif/yd"),
    "fabric_yy": ("fabric yy", "actual yy", "body yy", "total yy", "consumption"),
    "trim_yy": ("trim yy", "trim consumption"),
    "cm": ("cm", "cmt", "sewing cost", "cut make"),
    "trim": ("trim", "trim cost", "trim packaging", "packaging"),
    "fob": ("fob", "offer fob", "price", "unit price", "cost"),
    "ldp": ("ldp", "landed price"),
    "retail": ("retail", "retail px", "ticket price"),
    "profit": ("profit", "margin"),
    "vendor": ("vendor", "supplier", "mgf", "manufacturer"),
    "factory": ("factory", "fty", "mill"),
    "color": ("color", "colors", "colour", "colorway", "combo", "print name"),
    "season": ("season",),
    "division": ("division", "div"),
    "bm": ("bm", "outlet bm", "buy month"),
    "projection": ("projection", "projected units"),
    "moq_mcq": ("moq/mcq", "moq mcq", "moq", "mcq"),
    "commit": ("commit", "commit date"),
    "sy": ("sy", "sample yardage"),
    "ceo": ("ceo", "ceo sample", "ceo sample date"),
    "comments": ("comments", "comment", "remark", "remarks", "notes"),
    "submit_stage": ("submit stage", "stage", "submit type"),
    "submit_date": ("submit date", "s/o date", "so date", "send date", "date"),
    "greige_commit": ("greige commit",),
    "color_call_due": ("color call due",),
    "silo_call_due": ("silo call due",),
    "size_breaks_due": ("size breaks due", "size break due"),
    "fabric_ex_mill": ("fabric ex mill", "fabric ex-mill"),
    "fabric_in_factory": ("fabric in factory", "fabric in-factory"),
    "cut_start": ("cut start",),
    "gac_date": ("gac date", "gac"),
    "ih_date": ("ih date", "ih", "ldr"),
}

REQUIRED_FIELDS: dict[str, tuple[str, ...]] = {
    "submit_solid": ("style", "color", "submit_stage", "submit_date"),
    "submit_print": ("style", "color", "submit_stage", "submit_date"),
    "trim_submit": ("style", "description", "submit_stage", "submit_date"),
    "mail_dispatch_bulk": ("style", "color", "submit_stage", "submit_date"),
    "mail_dispatch_ldip": ("style", "color", "submit_stage", "submit_date"),
    "mail_dispatch_print": ("style", "color", "submit_stage", "submit_date"),
    "costing_sheet": ("style", "fabric", "fabric_yy", "cm", "trim", "fob"),
    "costing_recap": ("style", "fabric", "fabric_yy", "cm", "trim", "fob"),
    "ceo_recap": ("style", "description", "fabric", "color", "projection", "moq_mcq", "sy", "ceo", "sketch"),
    "tp_photo": ("style", "description", "sketch"),
    "tna": (
        "greige_commit",
        "color_call_due",
        "silo_call_due",
        "size_breaks_due",
        "fabric_ex_mill",
        "fabric_in_factory",
        "cut_start",
        "gac_date",
        "ih_date",
    ),
}


def autofill_workbook(workbook, artifact_type: str, source_data: dict[str, Any]) -> dict[str, Any]:
    records, source_paths = build_records(source_data)
    _apply_artifact_defaults(records, source_data, artifact_type)
    form_sheet_records = _prepare_form_sheets(workbook, artifact_type, records)
    required = REQUIRED_FIELDS.get(artifact_type, ("style",))
    tables = (
        []
        if artifact_type in {"submit_solid", "submit_print", "trim_submit"}
        else _table_candidates(workbook)
    )
    filled_cells: list[str] = []
    tbd_cells: list[str] = []
    inserted_images = 0

    for record in records:
        candidate = _best_table(tables, record)
        if candidate:
            result = _fill_table_record(candidate, record, required)
            filled_cells.extend(result["filled"])
            tbd_cells.extend(result["tbd"])
            inserted_images += result["images"]

    label_result = _fill_label_fields(
        workbook,
        artifact_type,
        records,
        required,
        tables,
        form_sheet_records,
    )
    filled_cells.extend(label_result["filled"])
    tbd_cells.extend(label_result["tbd"])
    inserted_images += label_result["images"]

    notes = _write_source_notes(
        workbook,
        artifact_type,
        records,
        required,
        source_paths,
        source_data,
    )
    summary = {
        "artifact_type": artifact_type,
        "styles": _record_styles(records),
        "required_fields": list(required),
        "filled_cells": sorted(set(filled_cells)),
        "tbd_cells": sorted(set(tbd_cells)),
        "inserted_images": inserted_images,
        "source_note_rows": notes,
        "source_image_candidates": _image_candidates(records),
    }
    return summary


def _apply_artifact_defaults(records: list[dict[str, Any]], source_data: dict[str, Any], artifact_type: str) -> None:
    stage = source_data.get("caseStage") or source_data.get("stage")
    if not stage:
        stage = {
            "submit_solid": "SOLID SUBMIT",
            "submit_print": "PRINT SUBMIT",
            "trim_submit": "TRIM SUBMIT",
            "mail_dispatch_bulk": "BULK SUBMIT",
            "mail_dispatch_ldip": "L/DIP SUBMIT",
            "mail_dispatch_print": "PRINT S/O SUBMIT",
        }.get(artifact_type)
    for record in records:
        if stage and _blank(record.get("submit_stage")):
            record["submit_stage"] = stage


def _prepare_form_sheets(workbook, artifact_type: str, records: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    if artifact_type not in {"submit_solid", "submit_print", "trim_submit"} or not workbook.worksheets:
        return {}
    eligible = [sheet for sheet in workbook.worksheets if "cache" not in sheet.title.lower()]
    if not eligible:
        return {}
    record_values = " ".join(
        str(value)
        for record in records
        for value in record.values()
        if not _blank(value)
    ).lower()

    def score(sheet) -> int:
        text = " ".join(
            str(cell.value)
            for row in sheet.iter_rows(
                min_row=1,
                max_row=min(sheet.max_row, 60),
                min_col=1,
                max_col=min(sheet.max_column, 30),
            )
            for cell in row
            if cell.value not in (None, "")
        ).lower()
        tokens = [token for token in re.split(r"[^0-9a-z]+", record_values) if len(token) >= 5]
        return sum(1 for token in set(tokens) if token in text)

    base = max(eligible, key=score)
    selected = [base]
    for index in range(1, max(1, len(records))):
        clone = workbook.copy_worksheet(base)
        clone.title = _safe_sheet_title(workbook, str(records[index].get("style") or f"FORM {index + 1}"))
        selected.append(clone)
    for sheet in list(eligible):
        if sheet not in selected:
            workbook.remove(sheet)
    mapping: dict[str, dict[str, Any]] = {}
    for index, sheet in enumerate(selected):
        record = records[min(index, len(records) - 1)] if records else {}
        mapping[sheet.title] = {
            **record,
            "__form_targets__": _clear_form_fields(sheet),
        }
        stage = str(record.get("submit_stage") or "").strip()
        if stage:
            for row in sheet.iter_rows(
                min_row=1,
                max_row=min(sheet.max_row, 8),
                min_col=1,
                max_col=min(sheet.max_column, 12),
            ):
                for cell in row:
                    if isinstance(cell.value, str) and "submit form" in cell.value.lower():
                        cell.value = f"{stage} FORM"
                        break
    return mapping


def _safe_sheet_title(workbook, raw: str) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", raw).strip()[:31] or "FORM"
    candidate = base
    counter = 2
    while candidate in workbook.sheetnames:
        suffix = f" {counter}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    return candidate


def _clear_form_fields(sheet) -> dict[str, str]:
    clearable = {
        "style", "core_style", "description", "fabric", "color", "season", "division",
        "vendor", "factory", "submit_date", "submit_stage", "comments", "projection",
    }
    targets: dict[str, str] = {}
    for row in range(1, min(sheet.max_row, 180) + 1):
        for column in range(1, min(sheet.max_column, 60) + 1):
            label = sheet.cell(row=row, column=column).value
            field = _canonical_field(label) if isinstance(label, str) else None
            if field not in clearable:
                continue
            target = _label_target(sheet, row, column)
            if target is not None and not (isinstance(target.value, str) and target.value.startswith("=")):
                targets.setdefault(field, target.coordinate)
                target.value = None
    return targets


def build_records(source_data: dict[str, Any]) -> tuple[list[dict[str, Any]], list[str]]:
    mappings = list(_mapping_records(source_data))
    common: dict[str, Any] = {}
    by_style: dict[str, dict[str, Any]] = {}
    source_paths: list[str] = []

    for mapping in mappings:
        extracted = _extract_mapping(mapping)
        styles = _styles_from_value(extracted.get("style"))
        if not styles:
            styles = _styles_from_value(_without_paths(mapping))
        for value in _path_values(mapping):
            if value not in source_paths:
                source_paths.append(value)
        for key, value in extracted.items():
            if key != "style" and not _blank(value) and key not in common:
                common[key] = value
        for style in styles:
            record = by_style.setdefault(style, {"style": style})
            for key, value in extracted.items():
                if key != "style" and not _blank(value) and key not in record:
                    record[key] = value

    explicit_styles: list[str] = []
    for key in ("styles", "style", "style_no", "style_number"):
        explicit_styles.extend(_styles_from_value(source_data.get(key)))
    explicit_styles.extend(_styles_from_value(source_data.get("caseTitle")))
    for item in source_data.get("businessKeys", []) if isinstance(source_data.get("businessKeys"), list) else []:
        if isinstance(item, dict) and "style" in str(item.get("kind", "")).lower():
            explicit_styles.extend(_styles_from_value(item.get("value")))
        elif isinstance(item, dict):
            canonical = _canonical_field(str(item.get("kind", "")))
            value = item.get("value")
            if canonical and not _blank(value):
                common.setdefault(canonical, value)

    title = str(source_data.get("caseTitle") or "")
    if "season" not in common:
        match = re.search(r"\b(?:SP|SM|HO|HR|FA|FW)\s*'?\s*\d{2}\b", title, re.IGNORECASE)
        if match:
            common["season"] = match.group(0).replace(" ", "").upper()
    if "division" not in common:
        for division in ("OUTLET", "HAVEN", "HWW", "TXT", "DRESS", "CORE", "FRONTLINE"):
            if division in title.upper():
                common["division"] = division
                break

    ordered_styles: list[str] = []
    for style in [*explicit_styles, *by_style.keys()]:
        if style not in ordered_styles:
            ordered_styles.append(style)
    if ordered_styles:
        records = []
        for style in ordered_styles[:30]:
            records.append({**common, **by_style.get(style, {}), "style": style})
    else:
        records = [common]
    return records, source_paths[:100]


def _mapping_records(value: Any) -> Iterable[dict[str, Any]]:
    if isinstance(value, dict):
        yield value
        for child in value.values():
            if isinstance(child, (dict, list, tuple)):
                yield from _mapping_records(child)
    elif isinstance(value, (list, tuple)):
        for child in value:
            yield from _mapping_records(child)


def _extract_mapping(mapping: dict[str, Any]) -> dict[str, Any]:
    extracted: dict[str, Any] = {}
    for raw_key, value in mapping.items():
        canonical = _canonical_field(str(raw_key))
        if canonical and not _blank(value):
            extracted.setdefault(canonical, _display_value(value))
    if "style" not in extracted:
        styles = _styles_from_value(_without_paths(mapping))
        if styles:
            extracted["style"] = styles
    return extracted


def _canonical_field(value: str) -> str | None:
    normalized = _normalize(value)
    if not normalized:
        return None
    matches: list[tuple[int, str]] = []
    for canonical, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            target = _normalize(alias)
            if normalized == target or (len(target) >= 4 and target in normalized):
                matches.append((len(target), canonical))
    return max(matches, default=(0, ""))[1] or None


def _looks_like_field_label(value: str) -> bool:
    normalized = _normalize(value)
    if not normalized:
        return False
    for aliases in FIELD_ALIASES.values():
        for alias in aliases:
            target = _normalize(alias)
            if normalized == target:
                return True
            if " " in target and normalized.startswith(f"{target} "):
                return True
    return False


def _normalize(value: Any) -> str:
    text = str(value or "").replace("\xa0", " ").lower()
    text = re.sub(r"[_\-/]+", " ", text)
    return re.sub(r"[^0-9a-z가-힣]+", " ", text).strip()


def _styles_from_value(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, dict):
        text = json.dumps(value, ensure_ascii=False, default=str)
    elif isinstance(value, (list, tuple, set)):
        text = " ".join(str(item) for item in value)
    else:
        text = str(value)
    return list(dict.fromkeys(STYLE_PATTERN.findall(text)))


def _path_values(value: Any) -> Iterable[str]:
    if isinstance(value, dict):
        for key, child in value.items():
            if isinstance(child, str) and ("path" in str(key).lower() or Path(child).suffix.lower() in IMAGE_SUFFIXES):
                path = Path(child).expanduser()
                if path.is_absolute():
                    yield str(path)
            elif isinstance(child, (dict, list, tuple)):
                yield from _path_values(child)
    elif isinstance(value, (list, tuple)):
        for child in value:
            yield from _path_values(child)


def _without_paths(value: Any) -> Any:
    """Strip file paths before free-text style scanning.

    ``STYLE_PATTERN`` accepts any nine consecutive digits, so a folder name,
    hash or export filename that happens to contain nine digits would be read
    as a business style. That invents a record, fills an extra row and repeats
    the sketch image in a customer-facing workbook.
    """
    if isinstance(value, dict):
        return {
            key: _without_paths(child)
            for key, child in value.items()
            if not _is_path_like(key, child)
        }
    if isinstance(value, (list, tuple)):
        return [
            _without_paths(child)
            for child in value
            if not _is_path_like("", child)
        ]
    return value


def _is_path_like(key: Any, value: Any) -> bool:
    if not isinstance(value, str):
        return False
    if "path" in str(key).lower():
        return True
    candidate = value.strip()
    if not candidate:
        return False
    if Path(candidate).suffix.lower() in IMAGE_SUFFIXES:
        return True
    return bool(re.match(r"^[A-Za-z]:[\\/]", candidate)) or candidate.startswith("\\\\")


def _display_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, int, float)) and not isinstance(value, bool):
        return value
    if isinstance(value, (list, tuple, set)):
        return "\n".join(str(item) for item in value if not _blank(item))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, default=str)[:2_000]
    return str(value).strip()


def _blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _table_candidates(workbook) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        if sheet.title == "SOURCE_NOTES":
            continue
        for row in range(1, min(sheet.max_row, 100) + 1):
            columns: dict[str, int] = {}
            for column in range(1, min(sheet.max_column, 60) + 1):
                value = sheet.cell(row=row, column=column).value
                if not isinstance(value, str):
                    continue
                canonical = _canonical_field(value)
                if canonical and canonical not in columns:
                    columns[canonical] = column
            if "style" in columns and len(columns) >= 2:
                candidates.append({"sheet": sheet, "header_row": row, "columns": columns})
    return candidates


def _best_table(tables: list[dict[str, Any]], record: dict[str, Any]) -> dict[str, Any] | None:
    if not tables:
        return None
    style = str(record.get("style") or "")
    division = _normalize(record.get("division"))

    def score(table: dict[str, Any]) -> int:
        sheet = table["sheet"]
        value = 0
        if style and _find_style_row(table, style):
            value += 100
        if division and division in _normalize(sheet.title):
            value += 30
        value += min(len(table["columns"]), 20)
        if "cache" in sheet.title.lower():
            value -= 100
        return value

    return max(tables, key=score)


def _find_style_row(table: dict[str, Any], style: str) -> int | None:
    sheet = table["sheet"]
    column = table["columns"]["style"]
    for row in range(table["header_row"] + 1, min(sheet.max_row + 1, table["header_row"] + 500)):
        if style and style in str(sheet.cell(row=row, column=column).value or ""):
            return row
    return None


def _fill_table_record(table: dict[str, Any], record: dict[str, Any], required: tuple[str, ...]) -> dict[str, Any]:
    sheet = table["sheet"]
    columns = table["columns"]
    style = str(record.get("style") or "")
    row = _find_style_row(table, style) if style else None
    if row is None:
        row = _blank_table_row(table)
    if row is None:
        row = sheet.max_row + 1
        _copy_row_presentation(sheet, max(table["header_row"] + 1, sheet.max_row), row)

    filled: list[str] = []
    tbd: list[str] = []
    images = 0
    for field, column in columns.items():
        cell = _top_left_cell(sheet, row, column)
        value = record.get(field)
        if field == "sketch":
            path = _image_path(value)
            if path and _insert_image(sheet, cell.coordinate, path):
                images += 1
                filled.append(f"{sheet.title}!{cell.coordinate}")
            elif field in required and _blank(cell.value):
                cell.value = "TBD"
                tbd.append(f"{sheet.title}!{cell.coordinate}")
            continue
        if _blank(cell.value):
            if not _blank(value):
                _set_cell_value(cell, value, field)
                filled.append(f"{sheet.title}!{cell.coordinate}")
            elif field in required:
                cell.value = "TBD"
                tbd.append(f"{sheet.title}!{cell.coordinate}")
    return {"filled": filled, "tbd": tbd, "images": images}


def _blank_table_row(table: dict[str, Any]) -> int | None:
    sheet = table["sheet"]
    style_column = table["columns"]["style"]
    start = table["header_row"] + 1
    for row in range(start, min(max(sheet.max_row + 15, start + 15), start + 500)):
        if _blank(sheet.cell(row=row, column=style_column).value):
            mapped_values = [sheet.cell(row=row, column=column).value for column in table["columns"].values()]
            if sum(not _blank(value) for value in mapped_values) <= 1:
                return row
    return None


def _copy_row_presentation(sheet, source_row: int, target_row: int) -> None:
    if source_row < 1 or target_row <= source_row:
        return
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, sheet.max_column + 1):
        source = sheet.cell(row=source_row, column=column)
        target = sheet.cell(row=target_row, column=column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)


def _fill_label_fields(
    workbook,
    artifact_type: str,
    records: list[dict[str, Any]],
    required: tuple[str, ...],
    tables: list[dict[str, Any]],
    sheet_records: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    header_coordinates = {
        (item["sheet"].title, item["header_row"], column)
        for item in tables
        for column in item["columns"].values()
    }
    filled: list[str] = []
    tbd: list[str] = []
    images = 0
    handled: set[tuple[str, str]] = set()
    tna_scope = _best_tna_scope(workbook, required) if artifact_type == "tna" else None
    sheets = [tna_scope[0]] if tna_scope else workbook.worksheets
    for sheet in sheets:
        if sheet.title == "SOURCE_NOTES":
            continue
        record = sheet_records.get(sheet.title, records[0] if records else {})
        min_row = tna_scope[1] if tna_scope and sheet is tna_scope[0] else 1
        max_row = tna_scope[2] if tna_scope and sheet is tna_scope[0] else min(sheet.max_row, 180)
        min_column = tna_scope[3] if tna_scope and sheet is tna_scope[0] else 1
        max_column = tna_scope[4] if tna_scope and sheet is tna_scope[0] else min(sheet.max_column, 60)
        for row in range(min_row, max_row + 1):
            for column in range(min_column, max_column + 1):
                if (sheet.title, row, column) in header_coordinates:
                    continue
                label = sheet.cell(row=row, column=column).value
                if not isinstance(label, str):
                    continue
                field = _canonical_field(label)
                if not field or (sheet.title, field) in handled:
                    continue
                form_target = record.get("__form_targets__", {}).get(field)
                target = sheet[form_target] if form_target else _label_target(sheet, row, column)
                if target is None or target.coordinate == sheet.cell(row=row, column=column).coordinate:
                    continue
                value = record.get(field)
                if field == "sketch":
                    path = _image_path(value)
                    if path and _insert_image(sheet, target.coordinate, path):
                        images += 1
                        filled.append(f"{sheet.title}!{target.coordinate}")
                        handled.add((sheet.title, field))
                    elif field in required and _blank(target.value):
                        target.value = "TBD"
                        tbd.append(f"{sheet.title}!{target.coordinate}")
                        handled.add((sheet.title, field))
                else:
                    replace_existing = artifact_type == "tna" and field in required and not _blank(value)
                    if not (_blank(target.value) or replace_existing):
                        continue
                    if not _blank(value):
                        _set_cell_value(target, value, field)
                        filled.append(f"{sheet.title}!{target.coordinate}")
                        handled.add((sheet.title, field))
                    elif field in required:
                        target.value = "TBD"
                        tbd.append(f"{sheet.title}!{target.coordinate}")
                        handled.add((sheet.title, field))
    return {"filled": filled, "tbd": tbd, "images": images}


def _best_tna_scope(workbook, required: tuple[str, ...]):
    best = None
    best_score = -1
    row_window = 10
    column_window = 10
    for sheet in workbook.worksheets:
        if sheet.title == "SOURCE_NOTES":
            continue
        max_row = min(sheet.max_row, 180)
        max_column = min(sheet.max_column, 60)
        for row_start in range(1, max_row + 1):
            row_end = min(max_row, row_start + row_window - 1)
            for column_start in range(1, max_column + 1):
                column_end = min(max_column, column_start + column_window - 1)
                found = set()
                for row in range(row_start, row_end + 1):
                    for column in range(column_start, column_end + 1):
                        value = sheet.cell(row=row, column=column).value
                        field = _canonical_field(value) if isinstance(value, str) else None
                        if field in required:
                            found.add(field)
                score = len(found) * 100 - (row_end - row_start) - (column_end - column_start)
                if score > best_score:
                    best = (sheet, row_start, row_end, column_start, column_end)
                    best_score = score
    return best


def _label_target(sheet, row: int, column: int):
    label_cell = _top_left_cell(sheet, row, column)
    merged = next(
        (
            item
            for item in sheet.merged_cells.ranges
            if item.min_row <= row <= item.max_row and item.min_col <= column <= item.max_col
        ),
        None,
    )
    min_column = merged.min_col if merged else column
    max_column = merged.max_col if merged else column
    max_row = merged.max_row if merged else row
    candidates = [
        *[(target_column, row) for target_column in range(max_column + 1, max_column + 6)],
        *[(min_column, target_row) for target_row in range(max_row + 1, max_row + 3)],
    ]
    available = []
    for target_column, target_row in candidates:
        if target_column > sheet.max_column + 2 or target_row > sheet.max_row + 2:
            continue
        target = _top_left_cell(sheet, target_row, target_column)
        if target.coordinate == label_cell.coordinate:
            continue
        if isinstance(target.value, str) and _looks_like_field_label(target.value):
            continue
        available.append(target)
    return next((cell for cell in available if not _blank(cell.value)), available[0] if available else None)


def _top_left_cell(sheet, row: int, column: int):
    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= column <= merged.max_col:
            return sheet.cell(row=merged.min_row, column=merged.min_col)
    return sheet.cell(row=row, column=column)


def _set_cell_value(cell, value: Any, field: str) -> None:
    cell.value = value
    if isinstance(value, (datetime, date)):
        cell.number_format = "yyyy-mm-dd"
    cell.alignment = copy(cell.alignment) if cell.alignment else Alignment()
    cell.alignment = Alignment(
        horizontal=cell.alignment.horizontal,
        vertical=cell.alignment.vertical or "top",
        text_rotation=cell.alignment.text_rotation,
        wrap_text=True,
        shrink_to_fit=cell.alignment.shrink_to_fit,
        indent=cell.alignment.indent,
    )
    cell.comment = Comment(f"ORBIT evidence-backed field: {field}", "HANSOLL ORBIT")


def _image_path(value: Any) -> Path | None:
    if isinstance(value, str):
        path = Path(value).expanduser()
        if path.is_absolute() and path.exists() and path.is_file() and path.suffix.lower() in IMAGE_SUFFIXES:
            return path
    return None


def _insert_image(sheet, coordinate: str, path: Path) -> bool:
    try:
        image = WorksheetImage(BytesIO(path.read_bytes()))
        scale = min(1.0, 110 / max(image.width, 1), 110 / max(image.height, 1))
        image.width = max(36, int(image.width * scale))
        image.height = max(36, int(image.height * scale))
        image.anchor = coordinate
        sheet.add_image(image)
        row = sheet[coordinate].row
        sheet.row_dimensions[row].height = max(sheet.row_dimensions[row].height or 15, image.height * 0.78)
        return True
    except (FileNotFoundError, OSError, ValueError):
        return False


def _write_source_notes(
    workbook,
    artifact_type: str,
    records: list[dict[str, Any]],
    required: tuple[str, ...],
    source_paths: list[str],
    source_data: dict[str, Any],
) -> int:
    if "SOURCE_NOTES" in workbook.sheetnames:
        del workbook["SOURCE_NOTES"]
    sheet = workbook.create_sheet("SOURCE_NOTES")
    sheet.sheet_properties.tabColor = "5B8C85"
    headers = ["STYLE", "FIELD", "VALUE", "STATUS", "SOURCE"]
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="355C58")
        cell.alignment = Alignment(vertical="center")
    row = 2
    fields = list(dict.fromkeys([*required, *sorted({key for record in records for key in record if key != "style"})]))
    source = "\n".join(source_paths[:8]) or "업무 건 및 연결된 검색 근거"
    for record in records or [{}]:
        style = str(record.get("style") or "TBD")
        for field in fields:
            value = record.get(field)
            status = (
                _field_provenance(field, value, source_data, artifact_type)
                if not _blank(value)
                else "TBD"
            )
            sheet.cell(row=row, column=1, value=style)
            sheet.cell(row=row, column=2, value=field)
            sheet.cell(row=row, column=3, value=_display_value(value) if not _blank(value) else "TBD")
            sheet.cell(row=row, column=4, value=status)
            sheet.cell(row=row, column=5, value=source)
            row += 1
    sheet.cell(row=row + 1, column=1, value="ARTIFACT TYPE")
    sheet.cell(row=row + 1, column=2, value=artifact_type)
    widths = {1: 16, 2: 24, 3: 48, 4: 14, 5: 70}
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
    for data_row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5):
        for cell in data_row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:E{max(1, row - 1)}"
    return max(0, row - 2)


def _field_provenance(
    field: str,
    value: Any,
    source_data: dict[str, Any],
    artifact_type: str,
) -> str:
    expected = _display_value(value)
    for raw_key, raw_value in source_data.items():
        if _canonical_field(str(raw_key)) == field and _display_value(raw_value) == expected:
            return "USER_INPUT"

    evidence = source_data.get("evidence")
    if isinstance(evidence, list):
        for item in evidence:
            if not isinstance(item, dict):
                continue
            candidate = _extract_mapping(item).get(field)
            if candidate is not None and _display_value(candidate) == expected:
                return "EVIDENCE"

    business_keys = source_data.get("businessKeys")
    if isinstance(business_keys, list):
        for item in business_keys:
            if not isinstance(item, dict):
                continue
            if _canonical_field(str(item.get("kind", ""))) != field:
                continue
            if _display_value(item.get("value")) == expected:
                return "CASE_DATA"

    if field == "submit_stage":
        case_stage = source_data.get("caseStage") or source_data.get("stage")
        if case_stage and _display_value(case_stage) == expected:
            return "CASE_DATA"
        if artifact_type.startswith("submit_") or artifact_type.startswith("mail_dispatch_"):
            return "RULE"

    return "INFERRED"


def _record_styles(records: list[dict[str, Any]]) -> list[str]:
    return list(dict.fromkeys(str(record["style"]) for record in records if not _blank(record.get("style"))))


def _image_candidates(records: list[dict[str, Any]]) -> list[str]:
    values: list[str] = []
    for record in records:
        path = _image_path(record.get("sketch"))
        if path and str(path) not in values:
            values.append(str(path))
    return values


def validate_autofill_contract(workbook, artifact_type: str, source_data: dict[str, Any], fill_summary: dict[str, Any]) -> list[dict[str, Any]]:
    required = REQUIRED_FIELDS.get(artifact_type, ("style",))
    findings: list[dict[str, Any]] = []
    notes = workbook["SOURCE_NOTES"] if "SOURCE_NOTES" in workbook.sheetnames else None
    findings.append({
        "ok": notes is not None and notes.max_row > 1,
        "code": "source_notes",
        "detail": "SOURCE_NOTES에 입력값·TBD·출처를 기록했습니다." if notes is not None and notes.max_row > 1 else "SOURCE_NOTES가 없거나 비어 있습니다.",
    })
    note_fields: dict[str, list[tuple[str, str]]] = {}
    if notes:
        for row in notes.iter_rows(min_row=2, values_only=True):
            field = str(row[1] or "")
            value = str(row[2] or "")
            status = str(row[3] or "")
            if field:
                note_fields.setdefault(field, []).append((value, status))
    supported_statuses = {"USER_INPUT", "EVIDENCE", "CASE_DATA", "RULE"}
    missing = [
        field
        for field in required
        if not any(
            value.strip() and value.strip().upper() != "TBD" and status in supported_statuses
            for value, status in note_fields.get(field, [])
        )
    ]
    findings.append({
        "ok": not missing,
        "code": "required_fields_supported",
        "detail": "필수 업무 항목에 확인 가능한 근거가 있습니다." if not missing else f"근거 확인이 필요한 필수 항목: {', '.join(missing)}",
        "blocking": False,
    })
    styles = [str(value) for value in fill_summary.get("styles", []) if str(value)]
    visible_text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        if sheet.title != "SOURCE_NOTES"
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    style_required = "style" in required
    missing_styles = [style for style in styles if style not in visible_text] if style_required else []
    findings.append({
        "ok": not style_required or not styles or not missing_styles,
        "code": "style_cells_filled",
        "detail": (
            "이 산출물은 스타일 번호가 필수 항목이 아닙니다."
            if not style_required
            else "요청 Style이 회사 양식 본문에 입력되었습니다."
            if styles and not missing_styles
            else "Style 근거가 없어 TBD로 남겼습니다."
            if not styles
            else f"본문에서 Style을 찾지 못했습니다: {', '.join(missing_styles)}"
        ),
        "blocking": False,
    })
    image_candidates = fill_summary.get("source_image_candidates", [])
    inserted_images = int(fill_summary.get("inserted_images") or 0)
    findings.append({
        "ok": not image_candidates or inserted_images > 0,
        "code": "sketch_inserted",
        "detail": f"확인된 Sketch/TP Photo {inserted_images}개를 삽입했습니다." if inserted_images else "사용 가능한 Sketch/TP Photo가 없어 TBD로 남겼습니다.",
        "blocking": False,
    })
    formula_errors = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and re.search(r"#(?:REF|DIV/0|VALUE|NAME|N/A)!?", cell.value, re.IGNORECASE):
                    formula_errors.append(f"{sheet.title}!{cell.coordinate}")
    findings.append({
        "ok": not formula_errors,
        "code": "formula_references",
        "detail": "수식에서 명백한 오류 참조를 찾지 못했습니다." if not formula_errors else f"오류 수식 참조: {', '.join(formula_errors[:8])}",
    })
    return findings
