from __future__ import annotations

import re
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


@dataclass(frozen=True)
class SbdFinding:
    ok: bool
    code: str
    detail: str


SIZE_AND_TOTAL_COLS = (
    6,
    7,
    8,
    9,
    10,
    12,
    13,
    14,
    15,
    16,
    17,
    19,
    20,
    21,
    22,
    23,
    24,
    26,
    27,
    28,
    29,
    30,
    31,
)
SUB_PO_COLS = (3, 11, 18, 25)


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalize_color(value: object) -> str:
    text = cell_text(value)
    if text.startswith("="):
        return ""
    return text


def is_valid_color(value: object) -> bool:
    text = normalize_color(value)
    return bool(text) and text.upper() not in {"0", "#REF!", "TOTAL", "G/TOTAL"}


def parse_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        if not match:
            return None
        try:
            return float(match.group(0))
        except ValueError:
            return None


def row_hidden(sheet, row: int) -> bool:
    return bool(sheet.row_dimensions[row].hidden)


def row_sum(sheet_values, row: int) -> float:
    ae_value = parse_number(sheet_values.cell(row=row, column=31).value)
    if ae_value is not None:
        return ae_value
    explicit_total = 0.0
    found_total = False
    for col in (10, 17, 24, 30):
        value = parse_number(sheet_values.cell(row=row, column=col).value)
        if value is not None:
            explicit_total += value
            found_total = True
    if found_total:
        return explicit_total
    total = 0.0
    for col in SIZE_AND_TOTAL_COLS:
        if col in (10, 17, 24, 30, 31):
            continue
        value = parse_number(sheet_values.cell(row=row, column=col).value)
        if value is not None:
            total += value
    return total


def find_print_bounds(sheet) -> tuple[int, int, int, int] | None:
    print_area = sheet.print_area
    if not print_area:
        return None
    area = str(print_area).split(",")[0]
    if "!" in area:
        area = area.split("!", 1)[1]
    area = area.replace("$", "")
    try:
        return range_boundaries(area)
    except ValueError:
        return None


def find_gtotal_row(sheet_formula) -> int | None:
    for row in range(1, sheet_formula.max_row + 1):
        if cell_text(sheet_formula.cell(row=row, column=1).value).upper() == "G/TOTAL":
            return row
    return None


def visible_data_colors(sheet_formula, sheet_values, gtotal_row: int) -> list[str]:
    colors: list[str] = []
    for row in range(6, gtotal_row):
        if row_hidden(sheet_formula, row):
            continue
        color = normalize_color(sheet_values.cell(row=row, column=2).value) or normalize_color(
            sheet_formula.cell(row=row, column=2).value
        )
        if is_valid_color(color) and row_sum(sheet_values, row) > 0:
            if color not in colors:
                colors.append(color)
    return colors


def gtotal_summary(
    sheet_formula, sheet_values, gtotal_row: int
) -> tuple[list[tuple[int, str, float]], int | None]:
    summary_rows: list[tuple[int, str, float]] = []
    grand_row: int | None = None
    for row in range(gtotal_row, min(sheet_formula.max_row, gtotal_row + 30) + 1):
        if row_hidden(sheet_formula, row):
            continue
        total = row_sum(sheet_values, row)
        color = normalize_color(sheet_values.cell(row=row, column=2).value) or normalize_color(
            sheet_formula.cell(row=row, column=2).value
        )
        if is_valid_color(color) and total > 0:
            summary_rows.append((row, color, total))
            continue
        if summary_rows and total > 0:
            grand_row = row
            break
    if grand_row is None and len(summary_rows) == 1:
        grand_row = summary_rows[0][0]
    return summary_rows, grand_row


def check_required_sheets(workbook_formula) -> list[SbdFinding]:
    findings: list[SbdFinding] = []
    for sheet in ("SBD-FL22", "Trim detail"):
        findings.append(SbdFinding(sheet in workbook_formula.sheetnames, "sheet_exists", sheet))
    return findings


def check_hidden_business_rows(
    sheet_formula, sheet_values, print_max_row: int | None
) -> list[SbdFinding]:
    findings: list[SbdFinding] = []
    max_row = print_max_row or sheet_formula.max_row
    bad_rows: list[str] = []
    for row in range(1, max_row + 1):
        if not row_hidden(sheet_formula, row):
            continue
        has_positive_total = row_sum(sheet_values, row) > 0
        if has_positive_total:
            bad_rows.append(str(row))
    detail = "none" if not bad_rows else "hidden business rows: " + ", ".join(bad_rows[:30])
    findings.append(SbdFinding(not bad_rows, "hidden_business_rows", detail))
    return findings


def check_po_columns(sheet_formula, sheet_values, max_row: int | None) -> list[SbdFinding]:
    findings: list[SbdFinding] = []
    max_scan = max_row or sheet_formula.max_row
    gtotal_row = find_gtotal_row(sheet_formula)
    if gtotal_row is not None:
        max_scan = min(max_scan, gtotal_row - 1)
    bad_master: list[str] = []
    missing_master: list[str] = []
    bad_sub: list[str] = []
    unresolved_sub: list[str] = []
    for row in range(6, max_scan + 1):
        if row_hidden(sheet_formula, row) or row_sum(sheet_values, row) <= 0:
            continue
        color = normalize_color(sheet_values.cell(row=row, column=2).value) or normalize_color(
            sheet_formula.cell(row=row, column=2).value
        )
        if not is_valid_color(color):
            continue
        master_text = cell_text(sheet_values.cell(row=row, column=1).value) or cell_text(
            sheet_formula.cell(row=row, column=1).value
        )
        if not master_text:
            missing_master.append(f"A{row}")
        elif not re.fullmatch(r"3016\d+", master_text):
            bad_master.append(f"A{row}={master_text}")
        for col in SUB_PO_COLS:
            formula_value = sheet_formula.cell(row=row, column=col).value
            value_text = cell_text(sheet_values.cell(row=row, column=col).value)
            formula_text = cell_text(formula_value)
            text = value_text or formula_text
            if not text:
                continue
            if formula_text.startswith("=") and not value_text:
                unresolved_sub.append(
                    f"{sheet_formula.cell(row=row, column=col).coordinate}={formula_text}"
                )
                continue
            if not re.fullmatch(r"(650|651)\d+", text):
                bad_sub.append(f"{sheet_formula.cell(row=row, column=col).coordinate}={text}")
    findings.append(
        SbdFinding(
            not missing_master and not bad_master,
            "master_po",
            "ok"
            if not missing_master and not bad_master
            else "; ".join(
                part
                for part in (
                    "missing Master PO cells: " + ", ".join(missing_master[:20])
                    if missing_master
                    else "",
                    "bad Master PO cells: " + ", ".join(bad_master[:20]) if bad_master else "",
                )
                if part
            ),
        )
    )
    findings.append(
        SbdFinding(
            not bad_sub and not unresolved_sub,
            "sub_po",
            "ok"
            if not bad_sub and not unresolved_sub
            else "; ".join(
                part
                for part in (
                    "bad Sub PO cells: " + ", ".join(bad_sub[:20]) if bad_sub else "",
                    "unresolved Sub PO formulas: " + ", ".join(unresolved_sub[:20])
                    if unresolved_sub
                    else "",
                )
                if part
            ),
        )
    )
    return findings


def check_gtotal_layout(
    sheet_formula,
    sheet_values,
    *,
    expected_total: int | None,
    print_max_row: int | None,
) -> list[SbdFinding]:
    findings: list[SbdFinding] = []
    gtotal_row = find_gtotal_row(sheet_formula)
    if gtotal_row is None:
        return [SbdFinding(False, "gtotal_row", "G/TOTAL row missing")]
    findings.append(SbdFinding(True, "gtotal_row", f"row {gtotal_row}"))

    data_colors = visible_data_colors(sheet_formula, sheet_values, gtotal_row)
    summary_rows, grand_row = gtotal_summary(sheet_formula, sheet_values, gtotal_row)
    summary_colors = [color for _, color, _ in summary_rows]
    findings.append(
        SbdFinding(
            summary_colors == data_colors,
            "gtotal_colors",
            f"data colors={data_colors}; summary colors={summary_colors}",
        )
    )
    if grand_row is None:
        findings.append(SbdFinding(False, "grand_total_row", "grand total row missing"))
        return findings

    summary_total = sum(total for _, _, total in summary_rows)
    grand_total = row_sum(sheet_values, grand_row)
    if len(summary_rows) == 1 and grand_row == summary_rows[0][0]:
        summary_total = grand_total
    findings.append(
        SbdFinding(
            round(summary_total) == round(grand_total),
            "grand_total_sum",
            f"summary={summary_total:g}; grand={grand_total:g}; grand row={grand_row}",
        )
    )
    if expected_total is not None:
        findings.append(
            SbdFinding(
                round(grand_total) == expected_total,
                "expected_total",
                f"expected={expected_total}; grand={grand_total:g}",
            )
        )

    formula = cell_text(sheet_formula["C3"].value).replace("$", "").upper()
    findings.append(
        SbdFinding(
            formula == f"=AE{grand_row}",
            "c3_formula",
            f"C3 formula={cell_text(sheet_formula['C3'].value)!r}; expected '=AE{grand_row}'",
        )
    )
    if print_max_row is not None:
        findings.append(
            SbdFinding(
                print_max_row >= grand_row,
                "print_area",
                f"print max row={print_max_row}; grand row={grand_row}",
            )
        )
    return findings


def check_style_and_trim(
    workbook_formula, workbook_values, style: str | None, expected_total: int | None
) -> list[SbdFinding]:
    findings: list[SbdFinding] = []
    if style:
        sbd = workbook_formula["SBD-FL22"]
        trim_values = workbook_values["Trim detail"]
        findings.append(
            SbdFinding(
                style in cell_text(sbd["A1"].value),
                "style_header",
                f"SBD-FL22!A1={cell_text(sbd['A1'].value)!r}; style={style}",
            )
        )
        findings.append(
            SbdFinding(
                cell_text(trim_values["B5"].value).replace(".0", "") == style,
                "trim_style",
                f"Trim detail!B5={cell_text(trim_values['B5'].value)!r}; style={style}",
            )
        )
    if expected_total is not None:
        trim_values = workbook_values["Trim detail"]
        total = parse_number(trim_values["B7"].value)
        findings.append(
            SbdFinding(
                total is not None and round(total) == expected_total,
                "trim_total",
                f"Trim detail!B7={total}; expected={expected_total}",
            )
        )
    return findings


def validate_sbd_workbook(
    workbook_path: Path,
    *,
    style: str | None = None,
    expected_total: int | None = None,
) -> list[SbdFinding]:
    workbook_formula = None
    workbook_values = None
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="wmf image format is not supported.*")
        workbook_formula = load_workbook(workbook_path, data_only=False)
        workbook_values = load_workbook(workbook_path, data_only=True)
    try:
        findings = check_required_sheets(workbook_formula)
        if not all(item.ok for item in findings):
            return findings
        sheet_formula = workbook_formula["SBD-FL22"]
        sheet_values = workbook_values["SBD-FL22"]
        print_bounds = find_print_bounds(sheet_formula)
        print_max_row = print_bounds[3] if print_bounds else None
        findings.extend(
            check_style_and_trim(workbook_formula, workbook_values, style, expected_total)
        )
        findings.extend(check_hidden_business_rows(sheet_formula, sheet_values, print_max_row))
        findings.extend(check_po_columns(sheet_formula, sheet_values, print_max_row))
        findings.extend(
            check_gtotal_layout(
                sheet_formula,
                sheet_values,
                expected_total=expected_total,
                print_max_row=print_max_row,
            )
        )
        return findings
    finally:
        if workbook_formula is not None:
            workbook_formula.close()
        if workbook_values is not None:
            workbook_values.close()


def findings_ok(findings: Iterable[SbdFinding]) -> bool:
    return all(item.ok for item in findings)
