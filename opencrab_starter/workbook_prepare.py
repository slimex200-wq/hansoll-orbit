from __future__ import annotations

import hashlib
import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.packaging.custom import StringProperty

from .artifact_autofill import autofill_workbook, validate_autofill_contract


DISPATCH_SHEETS = {
    "solid_bulk": "Solid bulk",
    "solid_dip": "Solid DIP",
    "print": "Print s.off",
}
DISPATCH_ARTIFACT_TYPES = {
    "solid_bulk": "mail_dispatch_bulk",
    "solid_dip": "mail_dispatch_ldip",
    "print": "mail_dispatch_print",
}

ORBIT_PROPERTY_PREFIX = "ORBIT_"


def prepare_artifact_workbook(
    source: Path,
    output: Path,
    artifact_type: str,
    source_data: dict[str, Any] | None = None,
    sheet_kind: str | None = None,
) -> dict[str, Any]:
    """Create an evidence-traceable copy without filling unsupported business values."""
    source = source.resolve()
    output = output.resolve()
    _validate_copy_paths(source, output)
    source_data = source_data if isinstance(source_data, dict) else {}
    suffix = source.suffix.lower()

    if suffix == ".xlsm" and _contains_vba(source):
        raise ValueError(
            "Macro-enabled workbooks cannot be copied automatically. Use a reviewed macro-free .xlsx source."
        )

    if artifact_type.startswith("mail_dispatch_"):
        if not sheet_kind:
            raise ValueError("Dispatch workbook preparation requires a sheet kind.")
        result = prepare_dispatch_workbook(source, output, sheet_kind, source_data=source_data)
    elif suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(
            source,
            keep_vba=suffix == ".xlsm",
            keep_links=False,
        )
        try:
            fill_summary = autofill_workbook(workbook, artifact_type, source_data)
            _write_orbit_properties(
                workbook,
                source,
                artifact_type,
                source_data,
                fill_summary=fill_summary,
            )
            output.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(output)
            result = {
                "source": str(source),
                "output": str(output),
                "sheet_count": len(workbook.sheetnames),
                "fill_summary": fill_summary,
            }
        finally:
            workbook.close()
    elif suffix in {".xls", ".xlsb"}:
        raise ValueError(
            "Legacy .xls and .xlsb workbooks cannot be copied automatically. "
            "Use a reviewed macro-free .xlsx source."
        )
    else:
        raise ValueError(f"Unsupported workbook type: {suffix}")

    reopened = validate_prepared_artifact(output, artifact_type)
    blocking_failures = [
        item for item in reopened
        if not item["ok"] and item.get("blocking", True)
    ]
    if blocking_failures:
        output.unlink(missing_ok=True)
        _manifest_path(output).unlink(missing_ok=True)
        detail = "; ".join(item["detail"] for item in blocking_failures)
        raise ValueError(f"Prepared workbook verification failed: {detail}")
    return {**result, "artifact_type": artifact_type, "verification": reopened}


def prepare_dispatch_workbook(
    source: Path,
    output: Path,
    sheet_kind: str,
    source_data: dict[str, Any] | None = None,
) -> dict[str, Any]:
    source = source.resolve()
    output = output.resolve()
    if source == output:
        raise ValueError("Source workbook cannot be overwritten.")
    if output.exists():
        raise FileExistsError(f"Output workbook already exists: {output}")
    if not source.exists() or not source.is_file():
        raise FileNotFoundError(f"Source workbook not found: {source}")
    if source.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Dispatch preparation supports .xlsx and .xlsm workbooks only.")
    if sheet_kind not in DISPATCH_SHEETS:
        raise ValueError(f"Unsupported dispatch sheet kind: {sheet_kind}")

    sheet_name = DISPATCH_SHEETS[sheet_kind]
    if source.suffix.lower() == ".xlsm" and _contains_vba(source):
        raise ValueError("Macro-enabled dispatch templates cannot be copied automatically.")
    workbook = load_workbook(
        source,
        keep_vba=source.suffix.lower() == ".xlsm",
        keep_links=False,
    )
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Required dispatch sheet not found: {sheet_name}")
        target = workbook[sheet_name]
        for sheet in list(workbook.worksheets):
            if sheet is not target:
                workbook.remove(sheet)

        if sheet_kind in {"solid_bulk", "solid_dip"}:
            cleared = _clear_solid_dispatch(target)
        else:
            cleared = _clear_print_dispatch(target)

        dispatch_artifact_type = DISPATCH_ARTIFACT_TYPES[sheet_kind]
        fill_summary = (
            autofill_workbook(workbook, dispatch_artifact_type, source_data or {})
            if source_data
            else {}
        )
        _write_orbit_properties(
            workbook,
            source,
            dispatch_artifact_type,
            source_data or {},
            fill_summary=fill_summary,
        )

        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
    finally:
        workbook.close()

    return {
        "source": str(source),
        "output": str(output),
        "sheet": sheet_name,
        "sheet_kind": sheet_kind,
        "sheet_count": 1,
        "cleared_cells": cleared,
        "fill_summary": fill_summary,
    }


def validate_prepared_artifact(output: Path, artifact_type: str) -> list[dict[str, Any]]:
    output = output.resolve()
    findings: list[dict[str, Any]] = [
        {
            "ok": output.exists() and output.is_file() and output.stat().st_size > 0,
            "code": "output_readable",
            "detail": "결과 파일이 존재하고 비어 있지 않습니다."
            if output.exists() and output.is_file() and output.stat().st_size > 0
            else "결과 파일이 없거나 비어 있습니다.",
        }
    ]
    if not findings[0]["ok"]:
        return findings

    suffix = output.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        try:
            workbook = load_workbook(output, read_only=False, data_only=False, keep_links=False)
            try:
                props = {
                    item.name: str(item.value)
                    for item in workbook.custom_doc_props
                    if item.name.startswith(ORBIT_PROPERTY_PREFIX)
                }
                findings.extend([
                    {
                        "ok": bool(workbook.sheetnames),
                        "code": "workbook_reopened",
                        "detail": f"저장 후 다시 열어 {len(workbook.sheetnames)}개 시트를 확인했습니다.",
                    },
                    {
                        "ok": props.get("ORBIT_ARTIFACT_TYPE") == artifact_type,
                        "code": "source_traceability",
                        "detail": "업무 종류와 원본 근거 정보가 파일 속성에 기록되었습니다."
                        if props.get("ORBIT_ARTIFACT_TYPE") == artifact_type
                        else "파일 속성에서 업무 근거를 확인하지 못했습니다.",
                    },
                    {
                        "ok": props.get("ORBIT_NO_SOURCE_NO_FILL") == "true",
                        "code": "no_source_no_fill",
                        "detail": "근거 없는 값은 자동 입력하지 않는 규칙이 기록되었습니다."
                        if props.get("ORBIT_NO_SOURCE_NO_FILL") == "true"
                        else "근거 없는 값의 미입력 규칙이 누락되었습니다.",
                    },
                ])
                fill_summary_text = props.get("ORBIT_FILL_SUMMARY", "")
                source_data_text = props.get("ORBIT_SOURCE_DATA", "")
                if fill_summary_text:
                    try:
                        fill_summary = json.loads(fill_summary_text)
                        source_data = json.loads(source_data_text) if source_data_text else {}
                        findings.extend(
                            validate_autofill_contract(
                                workbook,
                                artifact_type,
                                source_data if isinstance(source_data, dict) else {},
                                fill_summary if isinstance(fill_summary, dict) else {},
                            )
                        )
                    except (json.JSONDecodeError, TypeError, ValueError):
                        findings.append({
                            "ok": False,
                            "code": "autofill_contract",
                            "detail": "자동작성 검증 정보를 읽지 못했습니다.",
                        })
            finally:
                workbook.close()
        except (BadZipFile, KeyError, OSError, ValueError) as exc:
            findings.append({
                "ok": False,
                "code": "workbook_reopened",
                "detail": f"저장한 파일을 다시 열지 못했습니다: {type(exc).__name__}",
            })
    else:
        manifest = _manifest_path(output)
        findings.append({
            "ok": manifest.exists(),
            "code": "source_traceability",
            "detail": "구형 Excel 형식의 원본 근거 보조 파일을 확인했습니다."
            if manifest.exists()
            else "구형 Excel 형식의 원본 근거 보조 파일이 없습니다.",
        })
    return findings


def _validate_copy_paths(source: Path, output: Path) -> None:
    if source == output:
        raise ValueError("Source workbook cannot be overwritten.")
    if output.exists():
        raise FileExistsError(f"Output workbook already exists: {output}")
    if not source.exists() or not source.is_file():
        raise FileNotFoundError(f"Source workbook not found: {source}")


def _contains_vba(source: Path) -> bool:
    try:
        with ZipFile(source) as archive:
            return any(name.lower().endswith("vbaproject.bin") for name in archive.namelist())
    except BadZipFile:
        return False


def _source_digest(source: Path) -> str:
    digest = hashlib.sha256()
    with source.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _compact_source_data(source_data: dict[str, Any]) -> str:
    return _bounded_json(source_data)


def _bounded_json(value: Any, max_length: int = 8_000) -> str:
    text = json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)
    if len(text) <= max_length:
        return text

    if isinstance(value, dict):
        compact: dict[str, Any] = {}
        for key, item in value.items():
            if key in {"evidence", "filled_cells", "tbd_cells", "source_image_candidates"}:
                compact[key] = list(item[:20]) if isinstance(item, list) else item
                compact[f"{key}_count"] = len(item) if isinstance(item, list) else 1
            elif key in {
                "artifact_type", "styles", "required_fields", "inserted_images",
                "source_note_rows", "caseTitle", "caseStage", "businessKeys",
                "pendingDecisions", "sourceSheet", "preparationRule",
            }:
                compact[key] = item
        compact["truncated"] = True
        text = json.dumps(compact, ensure_ascii=False, separators=(",", ":"), default=str)
        if len(text) <= max_length:
            return text

    fallback = {"truncated": True, "summary": str(value)[: max_length - 40]}
    return json.dumps(fallback, ensure_ascii=False, separators=(",", ":"), default=str)


def _write_orbit_properties(
    workbook,
    source: Path,
    artifact_type: str,
    source_data: dict[str, Any],
    *,
    fill_summary: dict[str, Any] | None = None,
) -> None:
    values = {
        "ORBIT_ARTIFACT_TYPE": artifact_type,
        "ORBIT_SOURCE_FILE": source.name,
        "ORBIT_SOURCE_SHA256": _source_digest(source),
        "ORBIT_SOURCE_DATA": _compact_source_data(source_data),
        "ORBIT_NO_SOURCE_NO_FILL": "true",
        "ORBIT_PREPARED_AT": datetime.now(UTC).isoformat(),
        "ORBIT_FILL_SUMMARY": _bounded_json(fill_summary or {}),
    }
    existing = {item.name for item in workbook.custom_doc_props}
    for name, value in values.items():
        if name in existing:
            del workbook.custom_doc_props[name]
        workbook.custom_doc_props.append(StringProperty(name=name, value=value))


def _manifest_path(output: Path) -> Path:
    return output.with_name(f"{output.name}.orbit-source.json")


def _write_source_manifest(
    output: Path,
    source: Path,
    artifact_type: str,
    source_data: dict[str, Any],
) -> None:
    _manifest_path(output).write_text(
        json.dumps(
            {
                "artifact_type": artifact_type,
                "source_file": source.name,
                "source_sha256": _source_digest(source),
                "source_data": source_data,
                "no_source_no_fill": True,
                "prepared_at": datetime.now(UTC).isoformat(),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def _clear_solid_dispatch(sheet) -> int:
    cleared = 0
    for row in range(3, 14):
        cleared += _clear_value(sheet.cell(row=row, column=3))
    for row in range(16, sheet.max_row + 1):
        for column in range(2, 17):
            cleared += _clear_value(sheet.cell(row=row, column=column))
    return cleared


def _clear_print_dispatch(sheet) -> int:
    cleared = 0
    for coordinate in (
        "C2",
        "C3",
        "C4",
        "C5",
        "C6",
        "C8",
        "C11",
        "C13",
        "C14",
        "C15",
        "E2",
        "E3",
        "E4",
        "E5",
        "E6",
        "E7",
        "E8",
        "E9",
        "E10",
        "E11",
        "E12",
        "G6",
        "G7",
        "G8",
        "G9",
        "G10",
        "G11",
        "G12",
    ):
        cleared += _clear_value(sheet[coordinate])
    for row in range(18, sheet.max_row + 1):
        for column in range(2, 17):
            cleared += _clear_value(sheet.cell(row=row, column=column))
    return cleared


def _clear_value(cell) -> int:
    if cell.value is None:
        return 0
    cell.value = None
    return 1
