from __future__ import annotations

import argparse
import json
import posixpath
import re
import sys
from collections.abc import Mapping
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


@dataclass(frozen=True)
class Finding:
    ok: bool
    code: str
    detail: str


@dataclass(frozen=True)
class DrawingCounts:
    images: int = 0
    shapes: int = 0


OOXML_SHAPE_TAGS = {"sp", "grpSp", "cxnSp", "graphicFrame", "contentPart"}


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def has_style(side: object) -> bool:
    return bool(getattr(side, "style", None))


def check_sheet_exists(workbook, sheet_name: str) -> list[Finding]:
    if sheet_name in workbook.sheetnames:
        return [Finding(True, "sheet_exists", sheet_name)]
    return [Finding(False, "missing_sheet", sheet_name)]


def check_count(actual: int, expected: object, code: str, label: str) -> Finding:
    if isinstance(expected, bool):
        return Finding(False, "invalid_count_spec", f"{label}: boolean is not a count expectation")
    if isinstance(expected, int):
        ok = actual == expected
        expectation = f"exactly {expected}"
    elif isinstance(expected, Mapping):
        exact = expected.get("exact")
        minimum = expected.get("min")
        maximum = expected.get("max")
        supplied = [value is not None for value in (exact, minimum, maximum)]
        values = [value for value in (exact, minimum, maximum) if value is not None]
        if not any(supplied) or any(
            isinstance(value, bool) or not isinstance(value, int) for value in values
        ):
            return Finding(
                False,
                "invalid_count_spec",
                f"{label}: expected an integer or exact/min/max integer mapping, got {expected!r}",
            )
        ok = True
        parts: list[str] = []
        if exact is not None:
            ok = ok and actual == exact
            parts.append(f"exactly {exact}")
        if minimum is not None:
            ok = ok and actual >= minimum
            parts.append(f"at least {minimum}")
        if maximum is not None:
            ok = ok and actual <= maximum
            parts.append(f"at most {maximum}")
        expectation = " and ".join(parts)
    else:
        return Finding(
            False,
            "invalid_count_spec",
            f"{label}: expected an integer or exact/min/max mapping, got {expected!r}",
        )
    return Finding(ok, code, f"{label}: expected {expectation}, actual {actual}")


def check_required_values(sheet, required_values: dict[str, Any]) -> list[Finding]:
    findings: list[Finding] = []
    for cell, expected in required_values.items():
        actual = cell_text(sheet[cell].value)
        expected_text = cell_text(expected)
        findings.append(
            Finding(
                actual == expected_text,
                "required_value",
                f"{sheet.title}!{cell}: expected {expected_text!r}, actual {actual!r}",
            )
        )
    return findings


def check_non_empty(sheet, cells: list[str]) -> list[Finding]:
    findings: list[Finding] = []
    for cell in cells:
        value = cell_text(sheet[cell].value)
        findings.append(
            Finding(
                bool(value),
                "non_empty",
                f"{sheet.title}!{cell}: {'present' if value else 'missing'}",
            )
        )
    return findings


def formula_text(cell) -> str | None:
    value = cell.value
    if cell.data_type == "f" or (isinstance(value, str) and value.startswith("=")):
        return str(value)
    return None


def check_formula_cells(sheet, cells: list[str]) -> list[Finding]:
    findings: list[Finding] = []
    for coordinate in cells:
        formula = formula_text(sheet[coordinate])
        findings.append(
            Finding(
                formula is not None,
                "formula_present",
                f"{sheet.title}!{coordinate}: {formula!r}"
                if formula is not None
                else f"{sheet.title}!{coordinate}: formula missing",
            )
        )
    return findings


def check_required_formulas(sheet, required_formulas: dict[str, Any]) -> list[Finding]:
    findings: list[Finding] = []
    for coordinate, expected in required_formulas.items():
        actual = formula_text(sheet[coordinate])
        if isinstance(expected, str):
            ok = actual == expected
            expectation = repr(expected)
        elif isinstance(expected, Mapping) and isinstance(expected.get("regex"), str):
            pattern = expected["regex"]
            try:
                ok = actual is not None and re.search(pattern, actual) is not None
            except re.error as exc:
                findings.append(
                    Finding(
                        False,
                        "invalid_formula_spec",
                        f"{sheet.title}!{coordinate}: invalid regex {pattern!r}: {exc}",
                    )
                )
                continue
            expectation = f"regex {pattern!r}"
        elif isinstance(expected, Mapping) and isinstance(expected.get("equals"), str):
            value = expected["equals"]
            ok = actual == value
            expectation = repr(value)
        else:
            findings.append(
                Finding(
                    False,
                    "invalid_formula_spec",
                    f"{sheet.title}!{coordinate}: expected formula must be a string or equals/regex mapping",
                )
            )
            continue
        findings.append(
            Finding(
                ok,
                "required_formula",
                f"{sheet.title}!{coordinate}: expected {expectation}, actual {actual!r}",
            )
        )
    return findings


def check_merged_ranges(sheet, ranges: list[str]) -> list[Finding]:
    existing = {str(item) for item in sheet.merged_cells.ranges}
    findings: list[Finding] = []
    for expected in ranges:
        findings.append(
            Finding(
                expected in existing,
                "merged_range",
                f"{sheet.title}!{expected}: {'present' if expected in existing else 'missing'}",
            )
        )
    return findings


def check_bordered_range(sheet, range_text: str) -> Finding:
    min_col, min_row, max_col, max_row = range_boundaries(range_text)
    missing: list[str] = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            border = cell.border
            if row == min_row and not has_style(border.top):
                missing.append(f"{cell.coordinate}.top")
            if row == max_row and not has_style(border.bottom):
                missing.append(f"{cell.coordinate}.bottom")
            if col == min_col and not has_style(border.left):
                missing.append(f"{cell.coordinate}.left")
            if col == max_col and not has_style(border.right):
                missing.append(f"{cell.coordinate}.right")
    return Finding(
        not missing,
        "bordered_range",
        f"{sheet.title}!{range_text}: {'ok' if not missing else 'missing ' + ', '.join(missing[:20])}",
    )


def split_quoted_csv(value: str) -> list[str]:
    parts: list[str] = []
    start = 0
    in_quotes = False
    index = 0
    while index < len(value):
        char = value[index]
        if char == "'":
            if in_quotes and index + 1 < len(value) and value[index + 1] == "'":
                index += 1
            else:
                in_quotes = not in_quotes
        elif char == "," and not in_quotes:
            parts.append(value[start:index])
            start = index + 1
        index += 1
    parts.append(value[start:])
    return parts


def normalize_print_area(value: object) -> list[str]:
    text = cell_text(value)
    if not text:
        return []
    normalized: list[str] = []
    for item in split_quoted_csv(text):
        coordinate = item.rsplit("!", 1)[-1].strip().replace("$", "").upper()
        if coordinate:
            normalized.append(coordinate)
    return normalized


def print_area_expectation_ranges(value: object) -> list[str]:
    if isinstance(value, str):
        return normalize_print_area(value)
    if isinstance(value, list) and all(isinstance(item, str) for item in value):
        ranges: list[str] = []
        for item in value:
            ranges.extend(normalize_print_area(item))
        return ranges
    return []


def check_print_area(sheet, expected: object) -> Finding:
    actual = normalize_print_area(sheet.print_area)
    if isinstance(expected, (str, list)):
        wanted = print_area_expectation_ranges(expected)
        ok = bool(wanted) and actual == wanted
        expectation = f"exactly {wanted}"
    elif isinstance(expected, Mapping):
        checks: list[bool] = []
        descriptions: list[str] = []
        if "required" in expected:
            required = expected["required"]
            if not isinstance(required, bool):
                return Finding(
                    False, "invalid_print_area_spec", f"{sheet.title}: required must be boolean"
                )
            checks.append(bool(actual) if required else True)
            descriptions.append("non-empty" if required else "optional")
        if "equals" in expected:
            wanted = print_area_expectation_ranges(expected["equals"])
            if not wanted:
                return Finding(
                    False,
                    "invalid_print_area_spec",
                    f"{sheet.title}: invalid equals value {expected['equals']!r}",
                )
            checks.append(actual == wanted)
            descriptions.append(f"exactly {wanted}")
        if "contains" in expected:
            wanted = print_area_expectation_ranges(expected["contains"])
            if not wanted:
                return Finding(
                    False,
                    "invalid_print_area_spec",
                    f"{sheet.title}: invalid contains value {expected['contains']!r}",
                )
            checks.append(all(item in actual for item in wanted))
            descriptions.append(f"containing {wanted}")
        if "count" in expected:
            count_finding = check_count(
                len(actual),
                expected["count"],
                "print_area",
                f"{sheet.title} print-area range count",
            )
            if count_finding.code == "invalid_count_spec":
                return count_finding
            checks.append(count_finding.ok)
            descriptions.append(count_finding.detail.split(": ", 1)[-1])
        if not checks:
            return Finding(
                False,
                "invalid_print_area_spec",
                f"{sheet.title}: no supported print-area expectation",
            )
        ok = all(checks)
        expectation = " and ".join(descriptions)
    else:
        return Finding(
            False, "invalid_print_area_spec", f"{sheet.title}: invalid expectation {expected!r}"
        )
    return Finding(ok, "print_area", f"{sheet.title}: expected {expectation}, actual {actual}")


def relationship_map(archive: ZipFile, part_name: str) -> dict[str, tuple[str, str]]:
    try:
        root = ElementTree.fromstring(archive.read(part_name))
    except KeyError:
        return {}
    relationships: dict[str, tuple[str, str]] = {}
    for element in root:
        if element.tag.rsplit("}", 1)[-1] != "Relationship":
            continue
        relationship_id = element.attrib.get("Id")
        target = element.attrib.get("Target")
        relationship_type = element.attrib.get("Type", "")
        if relationship_id and target:
            relationships[relationship_id] = (target, relationship_type)
    return relationships


def resolve_part_name(base_part: str, target: str) -> str:
    clean_target = target.replace("\\", "/")
    if clean_target.startswith("/"):
        return clean_target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_part), clean_target))


def worksheet_relationship_part(sheet_part: str) -> str:
    return posixpath.join(
        posixpath.dirname(sheet_part),
        "_rels",
        f"{posixpath.basename(sheet_part)}.rels",
    )


def relationship_id(element: ElementTree.Element) -> str | None:
    for key, value in element.attrib.items():
        if key.startswith("{") and key.endswith("}id"):
            return value
    return None


def read_ooxml_drawing_counts(workbook_path: Path) -> dict[str, DrawingCounts]:
    counts: dict[str, DrawingCounts] = {}
    with ZipFile(workbook_path) as archive:
        workbook_part = "xl/workbook.xml"
        workbook_root = ElementTree.fromstring(archive.read(workbook_part))
        workbook_relationships = relationship_map(archive, "xl/_rels/workbook.xml.rels")
        for sheet_element in workbook_root.iter():
            if sheet_element.tag.rsplit("}", 1)[-1] != "sheet":
                continue
            sheet_name = sheet_element.attrib.get("name")
            relation_id = relationship_id(sheet_element)
            if not sheet_name or not relation_id or relation_id not in workbook_relationships:
                continue
            target, _ = workbook_relationships[relation_id]
            sheet_part = resolve_part_name(workbook_part, target)
            sheet_relationships = relationship_map(archive, worksheet_relationship_part(sheet_part))
            images = 0
            shapes = 0
            for drawing_target, drawing_type in sheet_relationships.values():
                if not drawing_type.endswith("/drawing"):
                    continue
                drawing_part = resolve_part_name(sheet_part, drawing_target)
                try:
                    drawing_root = ElementTree.fromstring(archive.read(drawing_part))
                except KeyError:
                    continue
                for element in drawing_root.iter():
                    local_name = element.tag.rsplit("}", 1)[-1]
                    if local_name == "pic":
                        images += 1
                    elif local_name in OOXML_SHAPE_TAGS:
                        shapes += 1
            counts[sheet_name] = DrawingCounts(images=images, shapes=shapes)
    return counts


def sheet_match_config(spec: dict[str, Any]) -> dict[str, Any]:
    nested = spec.get("match", {})
    if nested is None:
        nested = {}
    if not isinstance(nested, Mapping):
        raise ValueError("sheet match must be a mapping")
    match = dict(nested)
    for key in ("name", "name_regex", "anchors", "anchor_regex"):
        if key in spec and key not in match:
            match[key] = spec[key]
    if not any(key in match for key in ("name", "name_regex", "anchors", "anchor_regex")):
        raise ValueError("sheet spec requires name, name_regex, anchors, or anchor_regex")
    return match


def sheet_matches(sheet, match: dict[str, Any]) -> bool:
    name = match.get("name")
    if name is not None and sheet.title != name:
        return False
    name_regex = match.get("name_regex")
    if name_regex is not None:
        if not isinstance(name_regex, str):
            raise ValueError("name_regex must be a string")
        if re.search(name_regex, sheet.title) is None:
            return False
    anchors = match.get("anchors", {})
    if not isinstance(anchors, Mapping):
        raise ValueError("anchors must be a cell-to-value mapping")
    for coordinate, expected in anchors.items():
        if cell_text(sheet[str(coordinate)].value) != cell_text(expected):
            return False
    anchor_regex = match.get("anchor_regex", {})
    if not isinstance(anchor_regex, Mapping):
        raise ValueError("anchor_regex must be a cell-to-regex mapping")
    for coordinate, pattern in anchor_regex.items():
        if not isinstance(pattern, str):
            raise ValueError("anchor_regex patterns must be strings")
        if re.search(pattern, cell_text(sheet[str(coordinate)].value)) is None:
            return False
    return True


def describe_sheet_match(match: dict[str, Any]) -> str:
    parts: list[str] = []
    for key in ("name", "name_regex", "anchors", "anchor_regex"):
        if key in match:
            parts.append(f"{key}={match[key]!r}")
    return ", ".join(parts)


def check_sheet_contents(
    sheet, spec: dict[str, Any], drawing_counts: DrawingCounts
) -> list[Finding]:
    findings = check_required_values(sheet, spec.get("required_values", {}))
    findings.extend(check_non_empty(sheet, spec.get("non_empty", [])))
    findings.extend(check_merged_ranges(sheet, spec.get("merged_ranges", [])))
    for range_text in spec.get("bordered_ranges", []):
        findings.append(check_bordered_range(sheet, range_text))
    findings.extend(check_formula_cells(sheet, spec.get("formula_cells", [])))
    required_formulas = spec.get("required_formulas", spec.get("formulas", {}))
    findings.extend(check_required_formulas(sheet, required_formulas))
    if "print_area" in spec:
        findings.append(check_print_area(sheet, spec["print_area"]))
    if "image_count" in spec:
        findings.append(
            check_count(
                drawing_counts.images,
                spec["image_count"],
                "image_count",
                f"{sheet.title} OOXML image count",
            )
        )
    if "shape_count" in spec:
        findings.append(
            check_count(
                drawing_counts.shapes,
                spec["shape_count"],
                "shape_count",
                f"{sheet.title} OOXML shape count",
            )
        )
    return findings


def check_sheet(
    workbook,
    spec: dict[str, Any],
    drawing_counts: dict[str, DrawingCounts] | None = None,
) -> list[Finding]:
    drawing_counts = drawing_counts or {}
    try:
        match = sheet_match_config(spec)
        if "name" in match:
            findings = check_sheet_exists(workbook, match["name"])
            if not findings[0].ok:
                return findings
            candidates = [workbook[match["name"]]]
        else:
            findings = []
            candidates = list(workbook.worksheets)
        matching_sheets = [sheet for sheet in candidates if sheet_matches(sheet, match)]
    except (KeyError, TypeError, ValueError, re.error) as exc:
        return [Finding(False, "invalid_sheet_match", str(exc))]

    count_expectation = spec.get("match_count")
    if count_expectation is not None or "name" not in match:
        findings.append(
            check_count(
                len(matching_sheets),
                count_expectation if count_expectation is not None else {"min": 1},
                "sheet_match_count",
                f"sheet match ({describe_sheet_match(match)})",
            )
        )
    elif not matching_sheets:
        findings.append(
            Finding(
                False, "sheet_match", f"{match['name']}: anchor or regex criteria did not match"
            )
        )

    if not matching_sheets:
        return findings
    for sheet in matching_sheets:
        findings.extend(
            check_sheet_contents(sheet, spec, drawing_counts.get(sheet.title, DrawingCounts()))
        )
    return findings


def validate_workbook(workbook_path: Path, spec: dict[str, Any]) -> list[Finding]:
    drawing_requested = any(
        "image_count" in sheet_spec or "shape_count" in sheet_spec
        for sheet_spec in spec.get("sheets", [])
    )
    drawing_error: str | None = None
    try:
        drawing_counts = read_ooxml_drawing_counts(workbook_path) if drawing_requested else {}
    except (ElementTree.ParseError, KeyError, OSError) as exc:
        drawing_counts = {}
        drawing_error = f"{type(exc).__name__}: {exc}"

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        findings: list[Finding] = []
        workbook_spec = spec.get("workbook", {})
        if workbook_spec is None:
            workbook_spec = {}
        if not isinstance(workbook_spec, Mapping):
            findings.append(Finding(False, "invalid_workbook_spec", "workbook must be a mapping"))
            workbook_spec = {}
        sheet_count_expectation = spec.get("sheet_count", workbook_spec.get("sheet_count"))
        if sheet_count_expectation is not None:
            findings.append(
                check_count(
                    len(workbook.sheetnames),
                    sheet_count_expectation,
                    "sheet_count",
                    "workbook sheet count",
                )
            )
        if drawing_error is not None:
            findings.append(Finding(False, "drawing_inspection", drawing_error))
        for sheet_spec in spec.get("sheets", []):
            findings.extend(check_sheet(workbook, sheet_spec, drawing_counts))
        if not spec.get("sheets"):
            findings.append(Finding(False, "empty_spec", "spec contains no sheets"))
        return findings
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate generated Excel workbook layout.")
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--spec", required=True, type=Path)
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    spec = json.loads(args.spec.read_text(encoding="utf-8"))
    findings = validate_workbook(args.workbook, spec)
    ok = all(item.ok for item in findings)
    payload = [item.__dict__ for item in findings]
    if args.json:
        print(json.dumps(payload, ensure_ascii=False, indent=2))
    else:
        print("PASS" if ok else "FAIL")
        for item in findings:
            mark = "OK" if item.ok else "ERR"
            print(f"- {mark} {item.code}: {item.detail}")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
