from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import sqlite3
import sys
import warnings
import zipfile
import zlib
from contextlib import closing
from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from typing import Callable, Iterable
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps

warnings.filterwarnings("ignore", category=DeprecationWarning)
warnings.filterwarnings("ignore", category=Image.DecompressionBombWarning)


STYLE_RE = re.compile(r"(?<!\d)(\d{9}[A-Z]?)(?!\d)", re.IGNORECASE)
SUPPORTED_OFFICE_EXTENSIONS = {".xlsx", ".xlsm", ".pptx", ".docx", ".pdf"}
SUPPORTED_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp", ".tif", ".tiff"}
SUPPORTED_EXTENSIONS = SUPPORTED_OFFICE_EXTENSIONS | SUPPORTED_IMAGE_EXTENSIONS

MAX_TEXT_CELL_CHARS = 160
MAX_NEARBY_ROWS = 3
MAX_NEARBY_COLS = 80
VECTOR_SIZE = 16
PROJECTION_BINS = 32
ORIENTATION_BINS = 8
MAX_IMAGE_PIXELS = 40_000_000

Image.MAX_IMAGE_PIXELS = MAX_IMAGE_PIXELS


@dataclass(frozen=True)
class SketchRecord:
    source_path: Path
    relative_path: str
    top_folder: str
    extension: str
    location: str
    style_no: str | None
    nearby_text: str
    image_bytes: bytes
    source: str


@dataclass(frozen=True)
class ImageFeatures:
    sha256: str
    width: int
    height: int
    ink_density: float
    bbox: dict[str, float]
    vector: list[float]


def utc_now() -> str:
    return datetime.now(UTC).isoformat()


def normalize_text(value: object) -> str:
    text = str(value or "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def find_style(text: str) -> str | None:
    match = STYLE_RE.search(text)
    return match.group(1).upper() if match else None


def safe_image_from_bytes(image_bytes: bytes) -> Image.Image | None:
    try:
        image = Image.open(BytesIO(image_bytes))
        width, height = image.size
        if width * height > MAX_IMAGE_PIXELS:
            return None
        image.load()
        return image
    except Exception:
        return None


def white_background(image: Image.Image) -> Image.Image:
    image = ImageOps.exif_transpose(image)
    if image.mode in {"RGBA", "LA"}:
        background = Image.new("RGBA", image.size, "WHITE")
        background.alpha_composite(image.convert("RGBA"))
        return background.convert("RGB")
    if image.mode == "P" and "transparency" in image.info:
        return white_background(image.convert("RGBA"))
    return image.convert("RGB")


def compute_features(image_bytes: bytes) -> ImageFeatures | None:
    original = safe_image_from_bytes(image_bytes)
    if original is None:
        return None
    image = white_background(original)
    width, height = image.size
    if width < 60 or height < 60:
        return None

    gray = ImageOps.grayscale(image)
    gray = ImageOps.autocontrast(gray)
    small = ImageOps.contain(gray, (256, 256), method=Image.Resampling.LANCZOS)
    canvas = Image.new("L", (256, 256), 255)
    offset = ((256 - small.width) // 2, (256 - small.height) // 2)
    canvas.paste(small, offset)

    pixels = flattened_pixels(canvas)
    ink = [max(0.0, (245 - value) / 245.0) for value in pixels]
    strong = [value > 0.12 for value in ink]
    ink_count = sum(1 for value in strong if value)
    ink_density = ink_count / len(strong)
    if ink_density < 0.003 or ink_density > 0.75:
        return None

    xs = [index % 256 for index, value in enumerate(strong) if value]
    ys = [index // 256 for index, value in enumerate(strong) if value]
    min_x, max_x = min(xs), max(xs)
    min_y, max_y = min(ys), max(ys)
    bbox = {
        "x": round(min_x / 255, 6),
        "y": round(min_y / 255, 6),
        "w": round((max_x - min_x + 1) / 256, 6),
        "h": round((max_y - min_y + 1) / 256, 6),
    }

    tiny = canvas.resize((VECTOR_SIZE, VECTOR_SIZE), Image.Resampling.LANCZOS)
    tiny_ink = [max(0.0, (245 - value) / 245.0) for value in flattened_pixels(tiny)]

    horizontal: list[float] = []
    vertical: list[float] = []
    bin_size = 256 // PROJECTION_BINS
    for bin_index in range(PROJECTION_BINS):
        y0 = bin_index * bin_size
        y1 = 256 if bin_index == PROJECTION_BINS - 1 else y0 + bin_size
        total = 0.0
        for y in range(y0, y1):
            start = y * 256
            total += sum(ink[start : start + 256])
        horizontal.append(total / ((y1 - y0) * 256))
        x0 = bin_index * bin_size
        x1 = 256 if bin_index == PROJECTION_BINS - 1 else x0 + bin_size
        total = 0.0
        for y in range(256):
            row = y * 256
            total += sum(ink[row + x] for x in range(x0, x1))
        vertical.append(total / ((x1 - x0) * 256))

    orientation = [0.0] * ORIENTATION_BINS
    data = pixels
    for y in range(1, 255):
        row = y * 256
        for x in range(1, 255):
            gx = data[row + x + 1] - data[row + x - 1]
            gy = data[row + 256 + x] - data[row - 256 + x]
            mag = math.hypot(gx, gy)
            if mag < 18:
                continue
            angle = (math.atan2(gy, gx) + math.pi) / (2 * math.pi)
            orientation[min(ORIENTATION_BINS - 1, int(angle * ORIENTATION_BINS))] += mag
    orientation_total = sum(orientation) or 1.0
    orientation = [value / orientation_total for value in orientation]

    stats = [
        width / max(width, height),
        height / max(width, height),
        ink_density,
        bbox["x"],
        bbox["y"],
        bbox["w"],
        bbox["h"],
    ]
    vector = stats + tiny_ink + horizontal + vertical + orientation
    norm = math.sqrt(sum(value * value for value in vector)) or 1.0
    vector = [value / norm for value in vector]
    digest = hashlib.sha256(image_bytes).hexdigest()
    return ImageFeatures(digest, width, height, ink_density, bbox, vector)


def flattened_pixels(image: Image.Image) -> list[int]:
    getter = getattr(image, "get_flattened_data", None)
    if getter is not None:
        return list(getter())
    return list(image.getdata())


def pack_vector(vector: list[float]) -> bytes:
    text = json.dumps([round(value, 8) for value in vector], separators=(",", ":"))
    return zlib.compress(text.encode("ascii"))


def unpack_vector(blob: bytes) -> list[float]:
    return json.loads(zlib.decompress(blob).decode("ascii"))


def cosine(left: list[float], right: list[float]) -> float:
    if len(left) != len(right):
        return 0.0
    return sum(a * b for a, b in zip(left, right))


def make_thumb(image_bytes: bytes, out_path: Path, size: int = 192) -> None:
    image = safe_image_from_bytes(image_bytes)
    if image is None:
        return
    image = white_background(image)
    image.thumbnail((size, size), Image.Resampling.LANCZOS)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    image.save(out_path, format="JPEG", quality=72, optimize=True)


def row_nearby_text(ws, row: int, col: int) -> str:
    parts: list[str] = []
    min_row = max(1, row - MAX_NEARBY_ROWS)
    max_row = min(ws.max_row or row, row + MAX_NEARBY_ROWS)
    max_col = min(ws.max_column or col, MAX_NEARBY_COLS)
    for row_index in range(min_row, max_row + 1):
        row_parts: list[str] = []
        for col_index in range(1, max_col + 1):
            value = ws.cell(row_index, col_index).value
            text = normalize_text(value)
            if text:
                row_parts.append(text[:MAX_TEXT_CELL_CHARS])
        if row_parts:
            parts.append(" | ".join(row_parts))
    return " / ".join(parts)[:1200]


def extract_xlsx_images(path: Path, root: Path) -> list[SketchRecord]:
    records: list[SketchRecord] = []
    workbook = load_workbook(path, data_only=True, read_only=False)
    relative_path = str(path.relative_to(root))
    top_folder = Path(relative_path).parts[0] if Path(relative_path).parts else ""
    try:
        for ws in workbook.worksheets:
            for index, image in enumerate(getattr(ws, "_images", []), start=1):
                anchor = getattr(image, "anchor", None)
                marker = getattr(anchor, "_from", None)
                row = int(getattr(marker, "row", 0)) + 1
                col = int(getattr(marker, "col", 0)) + 1
                cell = f"{get_column_letter(max(col, 1))}{max(row, 1)}"
                try:
                    image_bytes = image._data()
                except Exception:
                    continue
                nearby_text = row_nearby_text(ws, max(row, 1), max(col, 1))
                style_no = find_style(nearby_text) or find_style(relative_path)
                records.append(
                    SketchRecord(
                        path,
                        relative_path,
                        top_folder,
                        path.suffix.lower(),
                        f"{ws.title}!{cell} image {index}",
                        style_no,
                        nearby_text,
                        image_bytes,
                        "xlsx_image",
                    )
                )
    finally:
        workbook.close()
    return records


def zip_texts(archive: zipfile.ZipFile, prefix: str, suffix: str) -> dict[str, str]:
    texts: dict[str, str] = {}
    for name in archive.namelist():
        if not name.startswith(prefix) or not name.endswith(suffix):
            continue
        try:
            root = ElementTree.fromstring(archive.read(name))
        except Exception:
            continue
        parts = [
            element.text.strip() for element in root.iter() if element.text and element.text.strip()
        ]
        texts[name] = normalize_text(" ".join(parts))[:1200]
    return texts


def extract_zip_media(
    path: Path, root: Path, prefix: str, text_prefix: str, text_suffix: str, source: str
) -> list[SketchRecord]:
    records: list[SketchRecord] = []
    relative_path = str(path.relative_to(root))
    top_folder = Path(relative_path).parts[0] if Path(relative_path).parts else ""
    with zipfile.ZipFile(path) as archive:
        texts = zip_texts(archive, text_prefix, text_suffix)
        joined_text = " / ".join(texts.values())[:1200]
        style_no = find_style(joined_text) or find_style(relative_path)
        media = sorted(
            name
            for name in archive.namelist()
            if name.startswith(prefix) and Path(name).suffix.lower() in SUPPORTED_IMAGE_EXTENSIONS
        )
        for index, name in enumerate(media, start=1):
            records.append(
                SketchRecord(
                    path,
                    relative_path,
                    top_folder,
                    path.suffix.lower(),
                    f"{name} image {index}",
                    style_no,
                    joined_text,
                    archive.read(name),
                    source,
                )
            )
    return records


def extract_pdf_images(path: Path, root: Path, max_pages: int) -> list[SketchRecord]:
    from pypdf import PdfReader

    records: list[SketchRecord] = []
    relative_path = str(path.relative_to(root))
    top_folder = Path(relative_path).parts[0] if Path(relative_path).parts else ""
    reader = PdfReader(str(path))
    for page_index, page in enumerate(reader.pages[:max_pages], start=1):
        try:
            text = normalize_text(page.extract_text() or "")[:1200]
        except Exception:
            text = ""
        style_no = find_style(text) or find_style(relative_path)
        for image_index, image_file in enumerate(getattr(page, "images", []), start=1):
            try:
                image_bytes = image_file.data
            except Exception:
                continue
            records.append(
                SketchRecord(
                    path,
                    relative_path,
                    top_folder,
                    ".pdf",
                    f"page {page_index} image {image_index}",
                    style_no,
                    text,
                    image_bytes,
                    "pdf_image",
                )
            )
    return records


def extract_image_file(path: Path, root: Path) -> list[SketchRecord]:
    relative_path = str(path.relative_to(root))
    top_folder = Path(relative_path).parts[0] if Path(relative_path).parts else ""
    return [
        SketchRecord(
            path,
            relative_path,
            top_folder,
            path.suffix.lower(),
            "image_file",
            find_style(relative_path),
            normalize_text(relative_path),
            path.read_bytes(),
            "image_file",
        )
    ]


def extract_records(
    path: Path, root: Path, max_pdf_pages: int
) -> tuple[str, list[SketchRecord], str | None]:
    suffix = path.suffix.lower()
    try:
        if suffix in {".xlsx", ".xlsm"}:
            return "parsed", extract_xlsx_images(path, root), None
        if suffix == ".pptx":
            return (
                "parsed",
                extract_zip_media(path, root, "ppt/media/", "ppt/slides/", ".xml", "pptx_media"),
                None,
            )
        if suffix == ".docx":
            return (
                "parsed",
                extract_zip_media(
                    path, root, "word/media/", "word/document.xml", ".xml", "docx_media"
                ),
                None,
            )
        if suffix == ".pdf":
            return "parsed", extract_pdf_images(path, root, max_pdf_pages), None
        if suffix in SUPPORTED_IMAGE_EXTENSIONS:
            return "parsed", extract_image_file(path, root), None
        return "unsupported", [], None
    except Exception as exc:
        return "error", [], f"{type(exc).__name__}: {exc}"


def init_db(db_path: Path) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with closing(sqlite3.connect(db_path)) as conn, conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS files (
                path TEXT PRIMARY KEY,
                relative_path TEXT NOT NULL,
                top_folder TEXT NOT NULL,
                extension TEXT NOT NULL,
                size INTEGER NOT NULL,
                mtime_ns INTEGER NOT NULL,
                parse_status TEXT NOT NULL,
                image_count INTEGER NOT NULL,
                error TEXT,
                indexed_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS sketches (
                sketch_id TEXT PRIMARY KEY,
                path TEXT NOT NULL,
                relative_path TEXT NOT NULL,
                top_folder TEXT NOT NULL,
                extension TEXT NOT NULL,
                location TEXT NOT NULL,
                style_no TEXT,
                nearby_text TEXT NOT NULL,
                image_sha256 TEXT NOT NULL,
                width INTEGER NOT NULL,
                height INTEGER NOT NULL,
                ink_density REAL NOT NULL,
                bbox_json TEXT NOT NULL,
                vector_dim INTEGER NOT NULL,
                vector_zlib BLOB NOT NULL,
                thumb_path TEXT,
                source TEXT NOT NULL,
                indexed_at TEXT NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_sketches_style ON sketches(style_no)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_sketches_path ON sketches(path)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_sketches_hash ON sketches(image_sha256)")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS ingest_runs (
                run_id TEXT PRIMARY KEY,
                started_at TEXT NOT NULL,
                completed_at TEXT,
                stats_json TEXT NOT NULL
            )
            """
        )


def iter_files(
    root: Path,
    include_tops: set[str],
    path_contains: list[str],
    *,
    on_error: Callable[[OSError], None] | None = None,
) -> Iterable[Path]:
    starts = [root / top for top in sorted(include_tops)] if include_tops else [root]
    lowered_terms = [term.lower() for term in path_contains if term]
    for start in starts:
        if not start.exists():
            continue
        for dirpath, dirnames, filenames in os.walk(start, onerror=on_error):
            dirnames[:] = [name for name in dirnames if not name.startswith(".")]
            for filename in filenames:
                if filename.startswith("~$"):
                    continue
                path = Path(dirpath) / filename
                if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
                    continue
                text_path = str(path).lower()
                if lowered_terms and not any(term in text_path for term in lowered_terms):
                    continue
                yield path


def prune_missing_files(
    conn: sqlite3.Connection,
    *,
    source_root: Path,
    active_tops: set[str] | None,
    path_contains: list[str],
    seen_paths: set[str],
) -> int:
    if active_tops is not None:
        if not active_tops:
            return 0
        placeholders = ",".join("?" for _ in active_tops)
        rows = conn.execute(
            f"SELECT path FROM files WHERE top_folder IN ({placeholders})",
            sorted(active_tops),
        ).fetchall()
    else:
        rows = conn.execute("SELECT path FROM files").fetchall()
    root_key = os.path.normcase(str(source_root.resolve()))
    seen_keys = {os.path.normcase(str(Path(path).resolve())) for path in seen_paths}
    lowered_terms = [term.strip().casefold() for term in path_contains if term.strip()]
    stale_paths: list[str] = []
    for (path,) in rows:
        path_key = os.path.normcase(str(Path(path).resolve()))
        try:
            if os.path.commonpath((root_key, path_key)) != root_key:
                continue
        except ValueError:
            continue
        if lowered_terms and not any(term in path_key.casefold() for term in lowered_terms):
            continue
        if path_key not in seen_keys:
            stale_paths.append(path)
    if not stale_paths:
        return 0
    conn.executemany("DELETE FROM sketches WHERE path = ?", ((path,) for path in stale_paths))
    conn.executemany("DELETE FROM files WHERE path = ?", ((path,) for path in stale_paths))
    return len(stale_paths)


def sketch_id_for(record: SketchRecord, features: ImageFeatures) -> str:
    material = f"{record.source_path}|{record.location}|{features.sha256}"
    return hashlib.sha256(material.encode("utf-8", "ignore")).hexdigest()[:24]


def index_file(
    conn: sqlite3.Connection,
    root: Path,
    path: Path,
    indexed_at: str,
    force: bool,
    thumb_dir: Path | None,
    max_pdf_pages: int,
) -> tuple[int, int, str, int]:
    relative_path = str(path.relative_to(root))
    top_folder = Path(relative_path).parts[0] if Path(relative_path).parts else ""
    suffix = path.suffix.lower()
    try:
        stat = path.stat()
    except OSError as exc:
        return 0, 1, f"stat_error: {exc}", 0
    existing = conn.execute(
        "SELECT size, mtime_ns, parse_status FROM files WHERE path = ?", (str(path),)
    ).fetchone()
    if (
        not force
        and existing
        and existing[0] == stat.st_size
        and existing[1] == stat.st_mtime_ns
        and existing[2] != "error"
    ):
        return 0, 0, "unchanged", 0

    status, records, error = extract_records(path, root, max_pdf_pages)
    conn.execute("DELETE FROM sketches WHERE path = ?", (str(path),))
    image_count = 0
    for record in records:
        features = compute_features(record.image_bytes)
        if features is None:
            continue
        sketch_id = sketch_id_for(record, features)
        thumb_path: str | None = None
        if thumb_dir is not None:
            thumb = thumb_dir / f"{features.sha256[:2]}" / f"{sketch_id}.jpg"
            make_thumb(record.image_bytes, thumb)
            thumb_path = str(thumb)
        conn.execute(
            """
            INSERT OR REPLACE INTO sketches (
                sketch_id, path, relative_path, top_folder, extension, location,
                style_no, nearby_text, image_sha256, width, height, ink_density,
                bbox_json, vector_dim, vector_zlib, thumb_path, source, indexed_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                sketch_id,
                str(path),
                record.relative_path,
                record.top_folder,
                record.extension,
                record.location,
                record.style_no,
                record.nearby_text,
                features.sha256,
                features.width,
                features.height,
                features.ink_density,
                json.dumps(features.bbox, separators=(",", ":")),
                len(features.vector),
                pack_vector(features.vector),
                thumb_path,
                record.source,
                indexed_at,
            ),
        )
        image_count += 1
    conn.execute(
        """
        INSERT INTO files (
            path, relative_path, top_folder, extension, size, mtime_ns,
            parse_status, image_count, error, indexed_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(path) DO UPDATE SET
            relative_path=excluded.relative_path,
            top_folder=excluded.top_folder,
            extension=excluded.extension,
            size=excluded.size,
            mtime_ns=excluded.mtime_ns,
            parse_status=excluded.parse_status,
            image_count=excluded.image_count,
            error=excluded.error,
            indexed_at=excluded.indexed_at
        """,
        (
            str(path),
            relative_path,
            top_folder,
            suffix,
            stat.st_size,
            stat.st_mtime_ns,
            status,
            image_count,
            error,
            indexed_at,
        ),
    )
    return 1, 0 if status != "error" else 1, status, image_count


def build_index(args: argparse.Namespace) -> int:
    root = args.root.expanduser().resolve()
    if not root.is_dir():
        raise FileNotFoundError(f"source root is missing or not a directory: {root}")
    include_tops = set(args.include_top or [])
    missing_tops = sorted(top for top in include_tops if not (root / top).is_dir())
    if missing_tops:
        raise FileNotFoundError(f"requested visual scope is missing: {', '.join(missing_tops)}")
    db_path = args.db.expanduser()
    if args.reset and db_path.exists():
        db_path.unlink()
    thumb_dir = args.thumb_dir.expanduser() if args.thumb_dir else None
    init_db(db_path)

    run_id = "visual-sketch-" + datetime.now(UTC).strftime("%Y%m%dT%H%M%S%fZ")
    started_at = utc_now()
    stats = {
        "root": str(root),
        "include_tops": sorted(args.include_top or []),
        "path_contains": args.path_contains or [],
        "max_files": args.max_files,
        "files_seen": 0,
        "files_indexed": 0,
        "files_error": 0,
        "unchanged": 0,
        "files_pruned": 0,
        "images_indexed": 0,
        "scan_errors": 0,
        "by_status": {},
    }
    with closing(sqlite3.connect(db_path)) as conn, conn:
        conn.execute(
            "INSERT OR REPLACE INTO ingest_runs(run_id, started_at, completed_at, stats_json) VALUES (?, ?, ?, ?)",
            (run_id, started_at, None, json.dumps(stats, ensure_ascii=False)),
        )
        seen_paths: set[str] = set()
        scan_errors: list[OSError] = []
        for path in iter_files(
            root,
            include_tops,
            args.path_contains or [],
            on_error=scan_errors.append,
        ):
            seen_paths.add(str(path))
            stats["files_seen"] += 1
            indexed, errored, status, image_count = index_file(
                conn,
                root,
                path,
                started_at,
                args.force,
                thumb_dir,
                args.max_pdf_pages,
            )
            stats["files_indexed"] += indexed
            stats["files_error"] += errored
            stats["images_indexed"] += image_count
            if status == "unchanged":
                stats["unchanged"] += 1
            stats["by_status"][status] = stats["by_status"].get(status, 0) + 1
            if args.max_files and stats["files_seen"] >= args.max_files:
                break
            if stats["files_seen"] % args.progress_every == 0:
                conn.commit()
                print(json.dumps(stats, ensure_ascii=False), flush=True)
        stats["scan_errors"] = len(scan_errors)
        if scan_errors:
            raise OSError(f"visual source scan was incomplete: {scan_errors[0]}")
        if not args.max_files and stats["files_seen"] > 0:
            active_tops = (
                {top for top in include_tops if (root / top).is_dir()} if include_tops else None
            )
            stats["files_pruned"] = prune_missing_files(
                conn,
                source_root=root,
                active_tops=active_tops,
                path_contains=args.path_contains or [],
                seen_paths=seen_paths,
            )
        completed_at = utc_now()
        conn.execute(
            "UPDATE ingest_runs SET completed_at = ?, stats_json = ? WHERE run_id = ?",
            (completed_at, json.dumps(stats, ensure_ascii=False), run_id),
        )
        conn.commit()
        total_sketches = conn.execute("SELECT COUNT(*) FROM sketches").fetchone()[0]
        total_styles = conn.execute(
            "SELECT COUNT(DISTINCT style_no) FROM sketches WHERE style_no IS NOT NULL"
        ).fetchone()[0]
    print(
        json.dumps(
            {
                "run_id": run_id,
                "total_sketches": total_sketches,
                "total_styles": total_styles,
                **stats,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


def search_index(args: argparse.Namespace) -> int:
    query_bytes = args.query_image.expanduser().read_bytes()
    features = compute_features(query_bytes)
    if features is None:
        print(json.dumps({"error": "query image could not be vectorized"}, ensure_ascii=False))
        return 2
    query_vector = features.vector
    with closing(sqlite3.connect(args.db.expanduser())) as conn, conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT sketch_id, relative_path, location, style_no, nearby_text, width, height,
                   ink_density, bbox_json, vector_zlib, thumb_path, source
            FROM sketches
            """
        ).fetchall()
    scored: list[dict[str, object]] = []
    for row in rows:
        vector = unpack_vector(row["vector_zlib"])
        score = cosine(query_vector, vector)
        scored.append(
            {
                "score": round(score, 6),
                "style_no": row["style_no"],
                "relative_path": row["relative_path"],
                "location": row["location"],
                "source": row["source"],
                "width": row["width"],
                "height": row["height"],
                "ink_density": round(row["ink_density"], 6),
                "bbox": json.loads(row["bbox_json"]),
                "thumb_path": row["thumb_path"],
                "nearby_text": row["nearby_text"][:500],
            }
        )
    scored.sort(key=lambda item: item["score"], reverse=True)
    print(json.dumps(scored[: args.limit], ensure_ascii=False, indent=2))
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build and query a compact visual sketch index.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    build = subparsers.add_parser("build")
    build.add_argument("--root", required=True, type=Path)
    build.add_argument("--db", default=Path("data/visual_sketch_index.sqlite"), type=Path)
    build.add_argument("--include-top", action="append")
    build.add_argument("--path-contains", action="append")
    build.add_argument("--thumb-dir", type=Path)
    build.add_argument("--force", action="store_true")
    build.add_argument("--reset", action="store_true")
    build.add_argument("--max-files", type=int)
    build.add_argument("--max-pdf-pages", type=int, default=6)
    build.add_argument("--progress-every", type=int, default=100)

    search = subparsers.add_parser("search")
    search.add_argument("--query-image", required=True, type=Path)
    search.add_argument("--db", default=Path("data/visual_sketch_index.sqlite"), type=Path)
    search.add_argument("--limit", type=int, default=10)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.command == "build":
        return build_index(args)
    if args.command == "search":
        return search_index(args)
    return 1


if __name__ == "__main__":
    sys.exit(main())
