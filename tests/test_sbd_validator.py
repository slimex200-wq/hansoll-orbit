from __future__ import annotations

import shutil
import unittest
import uuid
from pathlib import Path

from openpyxl import Workbook

from opencrab_starter.sbd_validator import findings_ok, validate_sbd_workbook


class SbdValidatorTests(unittest.TestCase):
    def test_validate_sbd_multicombo_passes(self) -> None:
        root = Path.cwd() / ".test_tmp" / f"sbd_validator_{uuid.uuid4().hex}"
        root.mkdir(parents=True)
        try:
            workbook_path = root / "264900039_sbd.xlsx"
            write_multicombo_workbook(workbook_path)

            findings = validate_sbd_workbook(
                workbook_path,
                style="264900039",
                expected_total=11775,
            )
        finally:
            shutil.rmtree(root, ignore_errors=True)

        self.assertTrue(findings_ok(findings), [item for item in findings if not item.ok])

    def test_validate_sbd_fails_hidden_business_row(self) -> None:
        root = Path.cwd() / ".test_tmp" / f"sbd_validator_{uuid.uuid4().hex}"
        root.mkdir(parents=True)
        try:
            workbook_path = root / "hidden_row.xlsx"
            write_multicombo_workbook(workbook_path, hidden_rows={6})

            findings = validate_sbd_workbook(workbook_path)
        finally:
            shutil.rmtree(root, ignore_errors=True)

        self.assertTrue(
            any(item.code == "hidden_business_rows" and not item.ok for item in findings)
        )

    def test_validate_sbd_fails_bad_c3_formula(self) -> None:
        root = Path.cwd() / ".test_tmp" / f"sbd_validator_{uuid.uuid4().hex}"
        root.mkdir(parents=True)
        try:
            workbook_path = root / "bad_c3.xlsx"
            write_multicombo_workbook(workbook_path, c3_formula="=AE19")

            findings = validate_sbd_workbook(workbook_path)
        finally:
            shutil.rmtree(root, ignore_errors=True)

        self.assertTrue(any(item.code == "c3_formula" and not item.ok for item in findings))

    def test_validate_sbd_fails_master_po_in_sub_po_column(self) -> None:
        root = Path.cwd() / ".test_tmp" / f"sbd_validator_{uuid.uuid4().hex}"
        root.mkdir(parents=True)
        try:
            workbook_path = root / "bad_sub_po.xlsx"
            write_multicombo_workbook(workbook_path, bad_sub_po=True)

            findings = validate_sbd_workbook(workbook_path)
        finally:
            shutil.rmtree(root, ignore_errors=True)

        self.assertTrue(any(item.code == "sub_po" and not item.ok for item in findings))

    def test_validate_sbd_fails_missing_master_po_on_business_row(self) -> None:
        root = Path.cwd() / ".test_tmp" / f"sbd_validator_{uuid.uuid4().hex}"
        root.mkdir(parents=True)
        try:
            workbook_path = root / "missing_master_po.xlsx"
            write_multicombo_workbook(workbook_path, master_po=None)

            findings = validate_sbd_workbook(workbook_path)
        finally:
            shutil.rmtree(root, ignore_errors=True)

        self.assertTrue(any(item.code == "master_po" and not item.ok for item in findings))

    def test_validate_sbd_fails_wrong_master_po_prefix(self) -> None:
        root = Path.cwd() / ".test_tmp" / f"sbd_validator_{uuid.uuid4().hex}"
        root.mkdir(parents=True)
        try:
            workbook_path = root / "bad_master_po.xlsx"
            write_multicombo_workbook(workbook_path, master_po="30192314")

            findings = validate_sbd_workbook(workbook_path)
        finally:
            shutil.rmtree(root, ignore_errors=True)

        self.assertTrue(any(item.code == "master_po" and not item.ok for item in findings))


def write_multicombo_workbook(
    workbook_path: Path,
    *,
    c3_formula: str = "=AE145",
    hidden_rows: set[int] | None = None,
    bad_sub_po: bool = False,
    master_po: str | None = "30162314",
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SBD-FL22"
    trim = workbook.create_sheet("Trim detail")

    sheet["A1"] = "S#264900039 ORDER RECAP"
    sheet["C3"] = c3_formula
    sheet.print_area = "$A$1:$AE$145"
    trim["B5"] = "264900039"
    trim["B7"] = "11,775 PCS"

    rows = [
        (6, "01 BLUE MAJESTY", 519),
        (7, "03 LIGHT SOFT JADE", 604),
        (8, "04 ENCHANTED BLUE", 942),
        (9, "20 RED POP", 902),
        (10, "40 BRILLIANT PURPLE", 663),
        (11, "80 IVY GREEN", 751),
        (13, "01 BLUE MAJESTY", 786),
        (14, "03 LIGHT SOFT JADE", 990),
        (15, "04 ENCHANTED BLUE", 1556),
        (16, "20 RED POP", 1502),
        (17, "40 BRILLIANT PURPLE", 1190),
        (18, "80 IVY GREEN", 1370),
    ]
    for row, color, total in rows:
        sheet.cell(row=row, column=1).value = master_po
        sheet.cell(row=row, column=2).value = color
        sheet.cell(row=row, column=31).value = total

    if bad_sub_po:
        sheet["C6"] = "30162314"

    summary = [
        (139, "01 BLUE MAJESTY", 1305),
        (140, "03 LIGHT SOFT JADE", 1594),
        (141, "04 ENCHANTED BLUE", 2498),
        (142, "20 RED POP", 2404),
        (143, "40 BRILLIANT PURPLE", 1853),
        (144, "80 IVY GREEN", 2121),
    ]
    for row, color, total in summary:
        sheet.cell(row=row, column=2).value = color
        sheet.cell(row=row, column=31).value = total
    sheet["A139"] = "G/TOTAL"
    sheet["AE145"] = 11775

    for row in hidden_rows or set():
        sheet.row_dimensions[row].hidden = True

    workbook.save(workbook_path)
    workbook.close()


if __name__ == "__main__":
    unittest.main()
