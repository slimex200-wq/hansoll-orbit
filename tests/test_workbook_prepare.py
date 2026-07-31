from __future__ import annotations

import shutil
import json
import unittest
import uuid
from pathlib import Path

from openpyxl import Workbook, load_workbook
from PIL import Image

from opencrab_starter.workbook_prepare import (
    prepare_artifact_workbook,
    prepare_dispatch_workbook,
    validate_prepared_artifact,
)


class WorkbookPrepareTests(unittest.TestCase):
    def setUp(self) -> None:
        self.root = Path.cwd() / ".test_tmp" / f"workbook_prepare_{uuid.uuid4().hex}"
        self.root.mkdir(parents=True)
        self.source = self.root / "dispatch.xlsx"
        workbook = Workbook()
        bulk = workbook.active
        bulk.title = "Solid bulk"
        bulk["A1"] = "Style#"
        bulk["B2"] = "Buying Agent/Agent Office"
        bulk["B3"] = "Season/Division"
        bulk["C2"] = "MGF Sourcing, Korea"
        bulk["C3"] = "OLD SEASON"
        bulk["B15"] = "Div"
        bulk["B16"] = "OUTLET"
        bulk["C16"] = "OLD SEASON"
        bulk["C3"].number_format = "@"

        dip = workbook.create_sheet("Solid DIP")
        dip["B2"] = "Buying Agent/Agent Office"
        dip["B3"] = "Season/Division"
        dip["C3"] = "OLD DIP"

        printed = workbook.create_sheet("Print s.off")
        printed["B2"] = "PRINT NAME:"
        printed["C2"] = "OLD PRINT"
        printed["D2"] = "SUBMIT DATE:"
        printed["E2"] = "OLD DATE"
        printed["B17"] = "DIVISION"
        printed["B18"] = "OUTLET"
        workbook.save(self.source)
        workbook.close()

    def tearDown(self) -> None:
        shutil.rmtree(self.root, ignore_errors=True)

    def test_prepares_single_blank_solid_bulk_sheet(self) -> None:
        output = self.root / "bulk.xlsx"
        result = prepare_dispatch_workbook(self.source, output, "solid_bulk")

        workbook = load_workbook(output)
        try:
            self.assertEqual(workbook.sheetnames, ["Solid bulk"])
            sheet = workbook["Solid bulk"]
            self.assertEqual(sheet["C2"].value, "MGF Sourcing, Korea")
            self.assertIsNone(sheet["C3"].value)
            self.assertIsNone(sheet["B16"].value)
            self.assertEqual(sheet["C3"].number_format, "@")
        finally:
            workbook.close()
        self.assertGreater(result["cleared_cells"], 0)

    def test_prepares_single_blank_print_sheet(self) -> None:
        output = self.root / "print.xlsx"
        prepare_dispatch_workbook(self.source, output, "print")

        workbook = load_workbook(output)
        try:
            self.assertEqual(workbook.sheetnames, ["Print s.off"])
            sheet = workbook["Print s.off"]
            self.assertEqual(sheet["B2"].value, "PRINT NAME:")
            self.assertEqual(sheet["D2"].value, "SUBMIT DATE:")
            self.assertIsNone(sheet["C2"].value)
            self.assertIsNone(sheet["E2"].value)
            self.assertIsNone(sheet["B18"].value)
        finally:
            workbook.close()

    def test_rejects_source_overwrite(self) -> None:
        with self.assertRaises(ValueError):
            prepare_dispatch_workbook(self.source, self.source, "solid_bulk")

    def test_rejects_existing_output_overwrite(self) -> None:
        output = self.root / "existing.xlsx"
        output.write_bytes(b"existing content")

        with self.assertRaises(FileExistsError):
            prepare_dispatch_workbook(self.source, output, "solid_bulk")
        self.assertEqual(output.read_bytes(), b"existing content")

    def test_generic_artifact_preserves_layout_and_records_source_trace(self) -> None:
        output = self.root / "costing.xlsx"
        result = prepare_artifact_workbook(
            self.source,
            output,
            "costing_sheet",
            {"styles": ["271900010"], "price": "TBD"},
        )

        self.assertEqual(result["sheet_count"], 4)
        workbook = load_workbook(output)
        try:
            self.assertEqual(
                workbook.sheetnames,
                ["Solid bulk", "Solid DIP", "Print s.off", "SOURCE_NOTES"],
            )
            properties = {item.name: str(item.value) for item in workbook.custom_doc_props}
            self.assertEqual(properties["ORBIT_ARTIFACT_TYPE"], "costing_sheet")
            self.assertEqual(properties["ORBIT_NO_SOURCE_NO_FILL"], "true")
            self.assertIn("271900010", properties["ORBIT_SOURCE_DATA"])
        finally:
            workbook.close()
        findings = validate_prepared_artifact(output, "costing_sheet")
        self.assertFalse(
            next(item for item in findings if item["code"] == "required_fields_supported")["ok"]
        )

    def test_generic_artifact_rejects_existing_output(self) -> None:
        output = self.root / "existing-generic.xlsx"
        output.write_bytes(b"existing content")
        with self.assertRaises(FileExistsError):
            prepare_artifact_workbook(self.source, output, "costing_sheet", {})

    def test_ceo_recap_fills_supported_cells_and_inserts_tp_photo(self) -> None:
        source = self.root / "ceo.xlsx"
        photo = self.root / "tp-photo.png"
        Image.new("RGB", (80, 80), "white").save(photo)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "APR TXT"
        headers = [
            "Outlet Style #",
            "TP Photos",
            "Style Description",
            "Fabric Information",
            "Colors",
            "Projection",
            "MOQ/MCQ",
            "SY",
            "CEO",
        ]
        for column, value in enumerate(headers, start=1):
            sheet.cell(row=1, column=column, value=value)
        workbook.save(source)
        workbook.close()

        output = self.root / "ceo-filled.xlsx"
        result = prepare_artifact_workbook(
            source,
            output,
            "ceo_recap",
            {
                "caseTitle": "SP27 OUTLET APR CEO Recap 271952230",
                "businessKeys": [
                    {"kind": "style", "value": "271952230"},
                    {"kind": "division", "value": "OUTLET"},
                ],
                "evidence": [{
                    "style_no": "271952230",
                    "style_description": "25 IN SS CREW",
                    "fabric_information": "95/5 Cotton Span Jersey",
                    "colors": "TA WHITE",
                    "projection": 3000,
                    "moq_mcq": "3000/500",
                    "sy": "7/10",
                    "ceo_sample_date": "7/18",
                    "image_path": str(photo),
                }],
            },
        )

        workbook = load_workbook(output)
        try:
            sheet = workbook["APR TXT"]
            self.assertEqual(sheet["A2"].value, "271952230")
            self.assertEqual(sheet["C2"].value, "25 IN SS CREW")
            self.assertEqual(sheet["F2"].value, 3000)
            self.assertEqual(len(sheet._images), 1)
            self.assertIn("SOURCE_NOTES", workbook.sheetnames)
        finally:
            workbook.close()
        self.assertEqual(result["fill_summary"]["inserted_images"], 1)
        self.assertTrue(all(item["ok"] for item in validate_prepared_artifact(output, "ceo_recap")))

    def test_costing_sheet_uses_tbd_for_missing_required_values(self) -> None:
        source = self.root / "costing-source.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Costing"
        for column, value in enumerate(
            ["Style#", "Fabric", "Actual YY", "CMT", "TRIM/PACKAGING", "FOB"],
            start=1,
        ):
            sheet.cell(row=1, column=column, value=value)
        workbook.save(source)
        workbook.close()

        output = self.root / "costing-filled.xlsx"
        prepare_artifact_workbook(
            source,
            output,
            "costing_sheet",
            {
                "styles": ["271900010"],
                "fabric": "FL-TA-CS-01",
                "actual_yy": 1.2,
                "cmt": 2.5,
                "trim_cost": 0.45,
            },
        )

        workbook = load_workbook(output)
        try:
            sheet = workbook["Costing"]
            self.assertEqual(sheet["A2"].value, "271900010")
            self.assertEqual(sheet["B2"].value, "FL-TA-CS-01")
            self.assertEqual(sheet["F2"].value, "TBD")
        finally:
            workbook.close()
        findings = validate_prepared_artifact(output, "costing_sheet")
        self.assertFalse(
            next(item for item in findings if item["code"] == "required_fields_supported")["ok"]
        )

    def test_tna_replaces_template_dates_and_allows_quality_level_schedule(self) -> None:
        source = self.root / "tna-source.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "1X1 RIB"
        labels = [
            ("Greige commit", "2025-07-03"),
            ("Color call due", "2025-07-17"),
            ("Silo call due", "2025-07-17"),
            ("Size breaks due", "2025-07-31"),
            ("Fabric Ex-mill", "2025-09-04"),
            ("Fabric In-factory", "2025-09-25"),
            ("Cut start", "2025-09-30"),
            ("GAC date", "2025-11-04"),
            ("IH date (LDR)", "2025-12-29"),
        ]
        for row, (label, old_value) in enumerate(labels, start=2):
            sheet.cell(row=row, column=1, value=label)
            sheet.cell(row=row, column=2, value=old_value)
        archive = workbook.create_sheet("ARCHIVE")
        archive["A2"] = "GAC date"
        archive["B2"] = "2021-01-01"
        archive["A3"] = "IH date"
        archive["B3"] = "2021-02-01"
        workbook.save(source)
        workbook.close()

        output = self.root / "tna-filled.xlsx"
        prepare_artifact_workbook(
            source,
            output,
            "tna",
            {
                "caseTitle": "SPRING 27 1X1 RIB TNA",
                "greige_commit": "2026-07-03",
                "color_call_due": "2026-07-17",
                "silo_call_due": "2026-07-17",
                "size_breaks_due": "2026-07-31",
                "fabric_ex_mill": "2026-09-04",
                "fabric_in_factory": "2026-09-25",
                "cut_start": "2026-09-30",
                "gac_date": "2026-11-04",
                "ih_date": "2026-12-29",
            },
        )

        workbook = load_workbook(output)
        try:
            sheet = workbook["1X1 RIB"]
            self.assertEqual(sheet["B2"].value, "2026-07-03")
            self.assertEqual(sheet["B10"].value, "2026-12-29")
            self.assertEqual(workbook["ARCHIVE"]["B2"].value, "2021-01-01")
        finally:
            workbook.close()
        self.assertTrue(all(item["ok"] for item in validate_prepared_artifact(output, "tna")))

    def test_solid_submit_fills_merged_form_fields_and_keeps_one_form(self) -> None:
        source = self.root / "solid-submit-source.xlsx"
        workbook = Workbook()
        first = workbook.active
        first.title = "OLD SAMPLE"
        first.merge_cells("A2:B2")
        first["A2"] = "Style Number"
        first.merge_cells("C2:D2")
        first["C2"] = "OLD STYLE"
        first.merge_cells("A3:B3")
        first["A3"] = "Color"
        first.merge_cells("C3:D3")
        first["C3"] = "OLD COLOR"
        first["A4"] = "Submit Date"
        first["B4"] = "OLD DATE"
        first["A5"] = "Submit Stage"
        first["B5"] = "OLD STAGE"
        other = workbook.create_sheet("OTHER SAMPLE")
        other["A1"] = "Style Number"
        other["B1"] = "OTHER"
        workbook.save(source)
        workbook.close()

        output = self.root / "solid-submit-filled.xlsx"
        prepare_artifact_workbook(
            source,
            output,
            "submit_solid",
            {
                "caseTitle": "SP27 OUTLET 271900010 Solid Submit",
                "style": "271900010",
                "color": "TA WHITE",
            },
        )

        workbook = load_workbook(output)
        try:
            self.assertEqual(workbook.sheetnames, ["OLD SAMPLE", "SOURCE_NOTES"])
            sheet = workbook["OLD SAMPLE"]
            self.assertEqual(sheet["C2"].value, "271900010")
            self.assertEqual(sheet["C3"].value, "TA WHITE")
            self.assertEqual(sheet["B5"].value, "SOLID SUBMIT")
            self.assertEqual(sheet["B4"].value, "TBD")
            notes = workbook["SOURCE_NOTES"]
            date_rows = [
                row for row in notes.iter_rows(min_row=2, values_only=True)
                if row[1] == "submit_date"
            ]
            self.assertTrue(date_rows)
            self.assertEqual(date_rows[0][3], "TBD")
        finally:
            workbook.close()
        findings = validate_prepared_artifact(output, "submit_solid")
        self.assertFalse(
            next(item for item in findings if item["code"] == "required_fields_supported")["ok"]
        )

    def test_solid_submit_accepts_an_explicit_submit_date_as_supported(self) -> None:
        source = self.root / "solid-submit-dated.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "FORM"
        sheet["A1"] = "Style Number"
        sheet["A2"] = "Color"
        sheet["A3"] = "Submit Date"
        sheet["A4"] = "Submit Stage"
        workbook.save(source)
        workbook.close()

        output = self.root / "solid-submit-dated-output.xlsx"
        prepare_artifact_workbook(
            source,
            output,
            "submit_solid",
            {
                "style": "271900010",
                "color": "TA WHITE",
                "submit_date": "2026-07-31",
            },
        )

        findings = validate_prepared_artifact(output, "submit_solid")
        self.assertTrue(
            next(item for item in findings if item["code"] == "required_fields_supported")["ok"]
        )

    def test_rejects_legacy_workbooks_from_automatic_copy(self) -> None:
        for suffix in (".xls", ".xlsb"):
            source = self.root / f"legacy{suffix}"
            source.write_bytes(b"legacy workbook")
            output = self.root / f"legacy-output{suffix}"
            with self.assertRaisesRegex(ValueError, "cannot be copied automatically"):
                prepare_artifact_workbook(source, output, "costing_sheet", {})

    def test_dispatch_uses_canonical_artifact_type_and_valid_json_metadata(self) -> None:
        output = self.root / "bulk-filled.xlsx"
        prepare_artifact_workbook(
            self.source,
            output,
            "mail_dispatch_bulk",
            {
                "style": "271900010",
                "color": "TA WHITE",
                "evidence": [{"body": "X" * 20_000}],
            },
            sheet_kind="solid_bulk",
        )

        workbook = load_workbook(output)
        try:
            properties = {item.name: str(item.value) for item in workbook.custom_doc_props}
            self.assertEqual(properties["ORBIT_ARTIFACT_TYPE"], "mail_dispatch_bulk")
            self.assertTrue(json.loads(properties["ORBIT_SOURCE_DATA"])["truncated"])
            self.assertEqual(
                json.loads(properties["ORBIT_FILL_SUMMARY"])["artifact_type"],
                "mail_dispatch_bulk",
            )
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
