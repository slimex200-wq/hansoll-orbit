from __future__ import annotations

import shutil
import unittest
import uuid
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as WorkbookImage
from openpyxl.styles import Border, Side
from PIL import Image

from scripts.validate_workbook_layout import validate_workbook


XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"


def inject_minimal_ooxml_shape(workbook_path: Path) -> None:
    with ZipFile(workbook_path) as archive:
        parts = {item.filename: archive.read(item.filename) for item in archive.infolist()}
    drawing_name = sorted(
        name for name in parts if name.startswith("xl/drawings/drawing") and name.endswith(".xml")
    )[0]
    drawing = ElementTree.fromstring(parts[drawing_name])
    shape = ElementTree.fromstring(
        f"""
        <xdr:twoCellAnchor xmlns:xdr="{XDR_NS}" xmlns:a="{A_NS}">
          <xdr:from>
            <xdr:col>3</xdr:col><xdr:colOff>0</xdr:colOff>
            <xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff>
          </xdr:from>
          <xdr:to>
            <xdr:col>4</xdr:col><xdr:colOff>0</xdr:colOff>
            <xdr:row>2</xdr:row><xdr:rowOff>0</xdr:rowOff>
          </xdr:to>
          <xdr:sp macro="" textlink="">
            <xdr:nvSpPr>
              <xdr:cNvPr id="999" name="Validation Shape"/>
              <xdr:cNvSpPr/>
            </xdr:nvSpPr>
            <xdr:spPr>
              <a:prstGeom prst="rect"><a:avLst/></a:prstGeom>
            </xdr:spPr>
            <xdr:txBody><a:bodyPr/><a:lstStyle/><a:p/></xdr:txBody>
          </xdr:sp>
          <xdr:clientData/>
        </xdr:twoCellAnchor>
        """
    )
    drawing.append(shape)
    parts[drawing_name] = ElementTree.tostring(drawing, encoding="utf-8", xml_declaration=True)

    replacement = workbook_path.with_name(f"{workbook_path.stem}.replacement.xlsx")
    with ZipFile(replacement, "w", compression=ZIP_DEFLATED) as archive:
        for name, content in parts.items():
            archive.writestr(name, content)
    replacement.replace(workbook_path)


class WorkbookLayoutValidationTests(unittest.TestCase):
    def test_validate_workbook_layout_passes_required_box(self) -> None:
        root = Path.cwd() / ".test_tmp" / f"workbook_layout_{uuid.uuid4().hex}"
        root.mkdir(parents=True)
        try:
            workbook_path = root / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Submit"
            sheet["A1"] = "PRINT SUBMIT FORM"
            sheet["A3"] = "STRIKE OFF SUBMIT"
            thin = Side(style="thin")
            for row in range(3, 5):
                for col in range(1, 4):
                    sheet.cell(row=row, column=col).border = Border(
                        top=thin if row == 3 else None,
                        bottom=thin if row == 4 else None,
                        left=thin if col == 1 else None,
                        right=thin if col == 3 else None,
                    )
            workbook.save(workbook_path)
            workbook.close()

            findings = validate_workbook(
                workbook_path,
                {
                    "sheets": [
                        {
                            "name": "Submit",
                            "required_values": {"A1": "PRINT SUBMIT FORM"},
                            "non_empty": ["A3"],
                            "bordered_ranges": ["A3:C4"],
                        }
                    ]
                },
            )
        finally:
            shutil.rmtree(root, ignore_errors=True)

        self.assertTrue(all(item.ok for item in findings))

    def test_validate_workbook_layout_fails_missing_border(self) -> None:
        root = Path.cwd() / ".test_tmp" / f"workbook_layout_{uuid.uuid4().hex}"
        root.mkdir(parents=True)
        try:
            workbook_path = root / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Submit"
            sheet["A1"] = "PRINT SUBMIT FORM"
            workbook.save(workbook_path)
            workbook.close()

            findings = validate_workbook(
                workbook_path,
                {"sheets": [{"name": "Submit", "bordered_ranges": ["A1:B2"]}]},
            )
        finally:
            shutil.rmtree(root, ignore_errors=True)

        self.assertFalse(all(item.ok for item in findings))
        self.assertTrue(any(item.code == "bordered_range" and not item.ok for item in findings))

    def test_validate_workbook_checks_dynamic_sheet_formula_print_area_and_drawings(self) -> None:
        root = Path.cwd() / ".test_tmp" / f"workbook_layout_{uuid.uuid4().hex}"
        root.mkdir(parents=True)
        try:
            workbook_path = root / "shape_fixture.xlsx"
            image_bytes = BytesIO()
            Image.new("RGB", (2, 2), color="red").save(image_bytes, format="PNG")
            image_bytes.seek(0)

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "264900039 BULK"
            sheet["A1"] = "PRINT SUBMIT FORM"
            sheet["C1"] = 1
            sheet["C2"] = 2
            sheet["C3"] = "=SUM(C1:C2)"
            sheet.print_area = "A1:I38"
            sheet.add_image(WorkbookImage(image_bytes), "B2")
            workbook.save(workbook_path)
            workbook.close()
            inject_minimal_ooxml_shape(workbook_path)

            spec = {
                "sheet_count": 1,
                "sheets": [
                    {
                        "match": {
                            "name_regex": r"^\d{9}\s+BULK$",
                            "anchors": {"A1": "PRINT SUBMIT FORM"},
                        },
                        "match_count": 1,
                        "print_area": "A1:I38",
                        "formula_cells": ["C3"],
                        "required_formulas": {"C3": {"regex": r"^=SUM\("}},
                        "image_count": 1,
                        "shape_count": {"min": 1},
                    }
                ],
            }
            findings = validate_workbook(workbook_path, spec)

            self.assertTrue(
                all(item.ok for item in findings), [item.detail for item in findings if not item.ok]
            )
            self.assertTrue(
                {
                    "sheet_count",
                    "sheet_match_count",
                    "print_area",
                    "formula_present",
                    "required_formula",
                    "image_count",
                    "shape_count",
                }.issubset({item.code for item in findings})
            )

            round_tripped = load_workbook(workbook_path)
            round_tripped.save(workbook_path)
            round_tripped.close()
            round_trip_findings = validate_workbook(workbook_path, spec)
            self.assertTrue(
                any(item.code == "shape_count" and not item.ok for item in round_trip_findings),
                "OOXML shape loss after an openpyxl round trip should be detected",
            )
        finally:
            shutil.rmtree(root, ignore_errors=True)

    def test_validate_workbook_reports_count_formula_print_area_and_drawing_failures(self) -> None:
        root = Path.cwd() / ".test_tmp" / f"workbook_layout_{uuid.uuid4().hex}"
        root.mkdir(parents=True)
        try:
            workbook_path = root / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Submit 271900001"
            sheet["A1"] = "PRINT SUBMIT FORM"
            sheet["C3"] = "not a formula"
            workbook.save(workbook_path)
            workbook.close()

            findings = validate_workbook(
                workbook_path,
                {
                    "sheet_count": 2,
                    "sheets": [
                        {
                            "anchors": {"A1": "PRINT SUBMIT FORM"},
                            "print_area": {"required": True},
                            "formula_cells": ["C3"],
                            "required_formulas": {"C3": "=SUM(C1:C2)"},
                            "image_count": {"min": 1},
                            "shape_count": {"min": 1},
                        }
                    ],
                },
            )
        finally:
            shutil.rmtree(root, ignore_errors=True)

        failed_codes = {item.code for item in findings if not item.ok}
        self.assertTrue(
            {
                "sheet_count",
                "print_area",
                "formula_present",
                "required_formula",
                "image_count",
                "shape_count",
            }.issubset(failed_codes)
        )


if __name__ == "__main__":
    unittest.main()
